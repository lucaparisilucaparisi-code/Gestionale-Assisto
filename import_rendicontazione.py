"""
Import delle rendicontazioni mensili da file Excel "Estratto Prefattura".

Il file ha un foglio per mese (es. "Estratto Settembre", titolo
"ESTRATTO SETTEMBRE 2025") con le colonne:
    Commessa | Scuola | Utenti | Totale ore lavorate in 100' | Pasti

La colonna "ore in 100'" corrisponde direttamente al valore che il
gestionale memorizza in rendicontazione.ore_lavorate_60 e su cui calcola
imponibile/IVA: va quindi inserita senza conversioni.

L'import abbina ogni riga a un UTENTE gia' presente in anagrafica
(per nome completo, con disambiguazione per scuola in caso di omonimi)
e non crea utenti nuovi.
"""

import re
import unicodedata

import openpyxl

# Nomi mese italiani -> numero
MESI_NOMI = {
    'gennaio': 1, 'febbraio': 2, 'marzo': 3, 'aprile': 4,
    'maggio': 5, 'giugno': 6, 'luglio': 7, 'agosto': 8,
    'settembre': 9, 'ottobre': 10, 'novembre': 11, 'dicembre': 12,
}


def normalizza(testo):
    """Minuscolo, senza accenti, spazi singoli. Per confronti robusti."""
    if testo is None:
        return ''
    s = str(testo).strip().lower()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = re.sub(r'\s+', ' ', s)
    return s


def chiave_nome(nome_completo):
    """Chiave di abbinamento indipendente dall'ordine dei token.
    'Mura Filippo' e 'Filippo Mura' producono la stessa chiave."""
    return tuple(sorted(normalizza(nome_completo).split()))


def parse_mese_anno(testo):
    """Estrae (mese, anno) da stringhe tipo 'ESTRATTO SETTEMBRE 2025'
    o 'Estratto Settembre'. Ritorna (mese|None, anno|None)."""
    t = normalizza(testo)
    mese = None
    for nome, num in MESI_NOMI.items():
        if re.search(r'\b' + nome + r'\b', t):
            mese = num
            break
    m = re.search(r'(20\d{2})', t)
    anno = int(m.group(1)) if m else None
    return mese, anno


def parse_ore(valore):
    """Converte un valore di ore in float decimale.
    Accetta numeri (45, 18.17) e stringhe ('1:30', '1,30'). None se vuoto."""
    if valore is None or valore == '':
        return None
    if isinstance(valore, bool):
        return None
    if isinstance(valore, (int, float)):
        return float(valore)
    s = str(valore).strip().replace(',', '.')
    if s == '':
        return None
    if ':' in s:
        parti = s.split(':')
        try:
            h = int(parti[0])
            mm = int(parti[1]) if len(parti) > 1 and parti[1] != '' else 0
            return h + mm / 60.0
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_pasti(valore):
    """Numero di pasti come intero. Vuoto/None -> 0."""
    if valore is None or valore == '':
        return 0
    try:
        return int(round(float(str(valore).replace(',', '.'))))
    except (ValueError, TypeError):
        return 0


def _trova_header(scanned_rows):
    """Cerca la riga di intestazione tra le prime righe del foglio.
    Ritorna (indice_0based, mappa_colonne) oppure (None, {})."""
    for i, riga in enumerate(scanned_rows):
        celle = [normalizza(c) if isinstance(c, str) else '' for c in riga]
        if 'commessa' in celle and any('utent' in c for c in celle):
            col = {}
            for j, c in enumerate(celle):
                if 'commessa' in c:
                    col['commessa'] = j
                elif 'utent' in c or 'nominativ' in c:
                    col['utente'] = j
                elif 'scuola' in c or 'plesso' in c or c == 'ic':
                    col.setdefault('scuola', j)
                elif 'ore' in c:
                    col.setdefault('ore', j)
                elif 'pasti' in c or 'past' in c:
                    col['pasti'] = j
            return i, col
    return None, {}


def parse_workbook(file_stream):
    """Legge il workbook e restituisce una lista di fogli.

    Ogni foglio: {
        'foglio': str, 'mese': int|None, 'anno': int|None,
        'riconosciuto': bool,
        'righe': [{'riga', 'nome_completo', 'commessa', 'scuola', 'ore', 'pasti'}]
    }
    """
    wb = openpyxl.load_workbook(file_stream, data_only=True, read_only=True)
    fogli = []
    try:
        for ws in wb.worksheets:
            mese, anno = parse_mese_anno(ws.title)

            scanned = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))

            # Mese/anno dal titolo nel contenuto (es. cella A1 "ESTRATTO SETTEMBRE 2025")
            for riga in scanned:
                for cella in riga:
                    if isinstance(cella, str) and 'estratto' in normalizza(cella):
                        m2, a2 = parse_mese_anno(cella)
                        if m2:
                            mese = m2
                        if a2:
                            anno = a2

            header_i, col = _trova_header(scanned)
            if header_i is None or 'utente' not in col or 'ore' not in col:
                fogli.append({'foglio': ws.title, 'mese': mese, 'anno': anno,
                              'riconosciuto': False, 'righe': []})
                continue

            righe = []
            header_row_1based = header_i + 1
            for ridx, r in enumerate(
                    ws.iter_rows(min_row=header_row_1based + 1, values_only=True),
                    start=header_row_1based + 1):
                def get(key):
                    j = col.get(key)
                    return r[j] if j is not None and j < len(r) else None

                utente = get('utente')
                commessa = get('commessa')

                if not utente or normalizza(utente) in ('', 'totale'):
                    continue
                if commessa is not None and normalizza(commessa) == 'totale':
                    continue

                righe.append({
                    'riga': ridx,
                    'nome_completo': str(utente).strip(),
                    'commessa': str(commessa).strip() if commessa else '',
                    'scuola': str(get('scuola')).strip() if get('scuola') else '',
                    'ore': parse_ore(get('ore')),
                    'pasti': parse_pasti(get('pasti')),
                })

            fogli.append({'foglio': ws.title, 'mese': mese, 'anno': anno,
                          'riconosciuto': True, 'righe': righe})
    finally:
        wb.close()
    return fogli


def build_index(utenti):
    """Costruisce {chiave_nome: [utente, ...]} dai record DB.
    utente: dict con almeno id, nome, cognome, scuola, commessa."""
    index = {}
    for u in utenti:
        k = chiave_nome(f"{u.get('cognome', '')} {u.get('nome', '')}")
        index.setdefault(k, []).append(u)
    return index


def _scuola_overlap(scuola_excel, scuola_db):
    """Quanti token significativi condividono le due scuole (per disambiguare omonimi)."""
    a = set(t for t in normalizza(scuola_excel).split() if len(t) > 2)
    b = set(t for t in normalizza(scuola_db).split() if len(t) > 2)
    return len(a & b)


def match_riga(riga, index):
    """Abbina una riga a un utente. Ritorna (utente|None, stato).
    stato: 'match' | 'non_trovato' | 'ambiguo'."""
    candidati = index.get(chiave_nome(riga['nome_completo']), [])
    if not candidati:
        return None, 'non_trovato'
    if len(candidati) == 1:
        return candidati[0], 'match'

    # Omonimi: disambigua per scuola
    punteggi = [(_scuola_overlap(riga.get('scuola', ''), c.get('scuola', '')), c)
                for c in candidati]
    punteggi.sort(key=lambda x: x[0], reverse=True)
    migliore, secondo = punteggi[0][0], (punteggi[1][0] if len(punteggi) > 1 else -1)
    if migliore > 0 and migliore > secondo:
        return punteggi[0][1], 'match'
    return None, 'ambiguo'


def analizza(fogli, utenti):
    """Abbina tutte le righe di tutti i fogli agli utenti.

    Ritorna una lista di fogli arricchiti con, per ogni riga, l'esito del
    match (utente_id, stato). Le righe senza valore ore valido sono marcate
    'senza_ore' e ignorate in scrittura.
    """
    index = build_index(utenti)
    risultati = []
    for foglio in fogli:
        match, non_trovati, ambigui, senza_ore = [], [], [], []
        for riga in foglio['righe']:
            if riga['ore'] is None:
                senza_ore.append(riga)
                continue
            utente, stato = match_riga(riga, index)
            voce = dict(riga)
            if stato == 'match':
                voce['utente_id'] = utente['id']
                voce['utente_nome'] = f"{utente.get('cognome', '')} {utente.get('nome', '')}".strip()
                voce['utente_scuola'] = utente.get('scuola', '')
                match.append(voce)
            elif stato == 'ambiguo':
                ambigui.append(voce)
            else:
                non_trovati.append(voce)
        risultati.append({
            'foglio': foglio['foglio'],
            'mese': foglio['mese'],
            'anno': foglio['anno'],
            'riconosciuto': foglio['riconosciuto'],
            'match': match,
            'non_trovati': non_trovati,
            'ambigui': ambigui,
            'senza_ore': senza_ore,
        })
    return risultati
