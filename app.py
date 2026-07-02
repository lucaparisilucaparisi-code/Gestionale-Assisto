#!/usr/bin/env python3
"""
Gestionale OEPAC - Sistema di Rendicontazione
"""

import os
import re
import csv
import hashlib
import secrets
from functools import wraps
from io import StringIO
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, make_response, session
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd

# WebAuthn (impronta digitale / Windows Hello)
try:
    from webauthn import (
        generate_registration_options,
        verify_registration_response,
        generate_authentication_options,
        verify_authentication_response,
        options_to_json,
    )
    from webauthn.helpers.structs import (
        PublicKeyCredentialDescriptor,
        AuthenticatorSelectionCriteria,
        UserVerificationRequirement,
        ResidentKeyRequirement,
    )
    from webauthn.helpers import base64url_to_bytes, bytes_to_base64url
    WEBAUTHN_AVAILABLE = True
except ImportError:
    WEBAUTHN_AVAILABLE = False

import config
import database as db
import import_rendicontazione
import import_dipendenti
from routes_export import export_bp
from routes_backup import backup_bp
from routes_migrazione import migrazione_bp
from routes_report_locale import report_locale_bp
from routes_utenti_dettaglio import utenti_dettaglio_bp
from routes_dashboard import dashboard_bp

logger = config.setup_logging()

def push_undo(action_type, data):
    """Salva un'azione nello stack undo persistente"""
    db.push_undo_action(action_type, data)


app = Flask(__name__)
app.register_blueprint(export_bp)
app.register_blueprint(backup_bp)
app.register_blueprint(migrazione_bp)
app.register_blueprint(report_locale_bp)
app.register_blueprint(utenti_dettaglio_bp)
app.register_blueprint(dashboard_bp)
app.config['UPLOAD_FOLDER'] = config.UPLOAD_FOLDER
app.config['EXPORT_FOLDER'] = config.EXPORT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = config.MAX_CONTENT_LENGTH
def _load_or_generate_secret_key():
    """Carica il secret_key da file, oppure lo genera random e lo persiste.
    Priorita': variabile ambiente FLASK_SECRET_KEY > file .flask_secret_key."""
    env_key = os.environ.get('FLASK_SECRET_KEY')
    if env_key:
        return env_key
    key_path = os.path.join(os.path.dirname(config.DATABASE_PATH), '.flask_secret_key')
    try:
        if os.path.exists(key_path):
            with open(key_path, 'rb') as f:
                data = f.read().strip()
                if len(data) >= 32:
                    return data
        # Non esiste o corrotto: genera e salva con permessi ristretti
        new_key = secrets.token_bytes(32)
        with open(key_path, 'wb') as f:
            f.write(new_key)
        try:
            os.chmod(key_path, 0o600)
        except OSError:
            pass  # Windows: chmod non applicabile
        return new_key
    except OSError as e:
        # Fallback SICURO: chiave casuale solo in memoria. Le sessioni si
        # invalideranno ai riavvii, ma la chiave non e' mai indovinabile.
        # Non derivare MAI il secret_key da un valore prevedibile (es. il path
        # del DB): renderebbe forgiabili i cookie di sessione firmati.
        logger.warning(f"Impossibile persistere secret_key ({e}), uso chiave casuale in memoria")
        return secrets.token_bytes(32)


app.secret_key = _load_or_generate_secret_key()

# Assicura che le cartelle esistano
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['EXPORT_FOLDER'], exist_ok=True)
os.makedirs(config.BACKUP_FOLDER, exist_ok=True)

# Configurazione sessione (sicurezza)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
# Secure e' opt-in: su http://localhost romperebbe il cookie; attivarlo solo
# quando si serve dietro HTTPS (OEPAC_HTTPS=1).
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('OEPAC_HTTPS', '0') == '1'
app.config['PERMANENT_SESSION_LIFETIME'] = 60 * 60 * 8  # 8 ore

# Inizializza database
db.init_db()


# ==================== AUTENTICAZIONE ====================
# Configurazione WebAuthn per localhost (single-user, single-device)
WEBAUTHN_RP_ID = 'localhost'
WEBAUTHN_RP_NAME = 'Assisto - Gestionale OEPAC'
WEBAUTHN_ORIGIN = 'http://localhost:5000'

# Route pubbliche che NON richiedono autenticazione
PUBLIC_ENDPOINTS = {
    'login_page',
    'setup_page',
    'api_auth_setup',
    'api_auth_login',
    'api_auth_status',
    'api_webauthn_auth_begin',
    'api_webauthn_auth_complete',
    'static',
}


@app.before_request
def require_authentication():
    """Protegge globalmente tutte le route tranne quelle in PUBLIC_ENDPOINTS."""
    endpoint = request.endpoint
    if endpoint is None:
        return  # 404 sara' gestito da Flask normalmente

    if endpoint in PUBLIC_ENDPOINTS:
        return

    # Se auth non configurata, vai a setup (tranne se gia' ci stai andando)
    if not db.auth_is_configured():
        if request.path.startswith('/api/'):
            return jsonify({'error': 'Sistema non configurato', 'code': 'SETUP_REQUIRED'}), 403
        return redirect(url_for('setup_page'))

    # Se non autenticato, blocca
    if not session.get('authenticated'):
        if request.path.startswith('/api/'):
            return jsonify({'error': 'Non autenticato', 'code': 'AUTH_REQUIRED'}), 401
        return redirect(url_for('login_page', next=request.path))


@app.errorhandler(404)
def handle_404(e):
    """Risposta coerente per risorsa non trovata (JSON per le API)."""
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Risorsa non trovata', 'code': 'NOT_FOUND'}), 404
    return "Pagina non trovata. <a href='/'>Torna alla home</a>", 404


@app.errorhandler(Exception)
def handle_unexpected(e):
    """Handler centralizzato per eccezioni non gestite.

    Logga i dettagli lato server e restituisce un messaggio generico al client,
    senza esporre str(e) (che potrebbe rivelare schema DB o percorsi). Le eccezioni
    HTTP esplicite (abort/404...) mantengono il loro codice."""
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e
    logger.error(f"Errore non gestito su {request.method} {request.path}: {e}", exc_info=True)
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Errore interno del server', 'code': 'INTERNAL_ERROR'}), 500
    return "Errore interno del server", 500


def login_required(f):
    """Decorator: richiede che l'utente sia autenticato.
    Se non configurato, reindirizza a /setup. Se non loggato, a /login."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not db.auth_is_configured():
            return redirect(url_for('setup_page'))
        if not session.get('authenticated'):
            # Se è una chiamata API, ritorna 401 JSON
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Non autenticato', 'code': 'AUTH_REQUIRED'}), 401
            # Altrimenti redirect a login
            return redirect(url_for('login_page', next=request.path))
        return f(*args, **kwargs)
    return decorated


def _login_user(metodo):
    """Imposta la sessione come autenticata."""
    # Rigenera il contesto di sessione al login (difesa anti session-fixation):
    # eventuali valori pre-login non vengono promossi a sessione autenticata.
    session.clear()
    session.permanent = True
    session['authenticated'] = True
    session['auth_method'] = metodo
    session['auth_time'] = datetime.now().isoformat()
    db.auth_record_login(metodo)


# ---------- PAGINE AUTH ----------

@app.route('/setup', methods=['GET'])
def setup_page():
    """Prima configurazione: crea username e password."""
    if db.auth_is_configured():
        return redirect(url_for('login_page'))
    return render_template('setup.html')


@app.route('/login', methods=['GET'])
def login_page():
    """Pagina di login."""
    if not db.auth_is_configured():
        return redirect(url_for('setup_page'))
    if session.get('authenticated'):
        return redirect(url_for('index'))
    return render_template('login.html',
                           webauthn_available=WEBAUTHN_AVAILABLE,
                           has_credentials=len(db.webauthn_get_credentials()) > 0)


@app.route('/api/auth/setup', methods=['POST'])
def api_auth_setup():
    """Crea l'account iniziale (solo al primo setup)."""
    if db.auth_is_configured():
        return jsonify({'error': 'Gia configurato'}), 400

    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''

    if len(username) < 3:
        return jsonify({'error': 'Username deve essere almeno 3 caratteri'}), 400
    if len(password) < 8:
        return jsonify({'error': 'Password deve essere almeno 8 caratteri'}), 400

    password_hash = generate_password_hash(password)
    db.auth_create_user(username, password_hash)
    _login_user('password')
    return jsonify({'success': True})


@app.route('/api/auth/login', methods=['POST'])
def api_auth_login():
    """Login con username e password."""
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''

    user = db.auth_get_user()
    if not user or user['username'] != username:
        return jsonify({'error': 'Credenziali non valide'}), 401
    if not check_password_hash(user['password_hash'], password):
        return jsonify({'error': 'Credenziali non valide'}), 401

    _login_user('password')
    return jsonify({'success': True})


@app.route('/api/auth/logout', methods=['POST'])
def api_auth_logout():
    """Logout: pulisce la sessione."""
    session.clear()
    return jsonify({'success': True})


@app.route('/api/auth/status')
def api_auth_status():
    """Ritorna lo stato di autenticazione corrente."""
    user = db.auth_get_user() if db.auth_is_configured() else None
    return jsonify({
        'configured': db.auth_is_configured(),
        'authenticated': bool(session.get('authenticated')),
        'method': session.get('auth_method'),
        'username': user['username'] if user else None,
        'webauthn_available': WEBAUTHN_AVAILABLE,
        'webauthn_registered': len(db.webauthn_get_credentials()) > 0 if db.auth_is_configured() else False,
    })


# ---------- WEBAUTHN (impronta digitale / Windows Hello) ----------

def _webauthn_user_id():
    """User handle WebAuthn: deterministico (single-user).
    Deve essere un byte string stabile ma non deducibile dall'username."""
    user = db.auth_get_user()
    if not user:
        return None
    # Deriva da data_creazione + username (stabile, 32 byte)
    seed = (user['username'] + user['data_creazione']).encode('utf-8')
    return hashlib.sha256(seed).digest()


@app.route('/api/auth/webauthn/register/begin', methods=['POST'])
@login_required
def api_webauthn_register_begin():
    """Step 1: genera le options per la registrazione di una nuova credenziale."""
    if not WEBAUTHN_AVAILABLE:
        return jsonify({'error': 'WebAuthn non disponibile (pacchetto non installato)'}), 500

    user = db.auth_get_user()
    if not user:
        return jsonify({'error': 'Utente non trovato'}), 400

    # Escludi credenziali gia' registrate
    existing = db.webauthn_get_credentials()
    exclude = [
        PublicKeyCredentialDescriptor(id=bytes(c['credential_id']))
        for c in existing
    ]

    options = generate_registration_options(
        rp_id=WEBAUTHN_RP_ID,
        rp_name=WEBAUTHN_RP_NAME,
        user_id=_webauthn_user_id(),
        user_name=user['username'],
        user_display_name=user['username'],
        exclude_credentials=exclude,
        authenticator_selection=AuthenticatorSelectionCriteria(
            user_verification=UserVerificationRequirement.REQUIRED,
            resident_key=ResidentKeyRequirement.PREFERRED,
        ),
    )

    # Salva challenge in sessione (sara' verificato nello step successivo)
    session['webauthn_challenge'] = bytes_to_base64url(options.challenge)

    return app.response_class(options_to_json(options), mimetype='application/json')


@app.route('/api/auth/webauthn/register/complete', methods=['POST'])
@login_required
def api_webauthn_register_complete():
    """Step 2: verifica la risposta del browser e salva la credenziale."""
    if not WEBAUTHN_AVAILABLE:
        return jsonify({'error': 'WebAuthn non disponibile'}), 500

    challenge_b64 = session.pop('webauthn_challenge', None)
    if not challenge_b64:
        return jsonify({'error': 'Challenge non trovata o scaduta'}), 400

    data = request.get_json() or {}
    nome = (data.get('nome') or 'Impronta').strip()[:50]
    credential = data.get('credential')
    if not credential:
        return jsonify({'error': 'Credential mancante'}), 400

    try:
        verification = verify_registration_response(
            credential=credential,
            expected_challenge=base64url_to_bytes(challenge_b64),
            expected_origin=WEBAUTHN_ORIGIN,
            expected_rp_id=WEBAUTHN_RP_ID,
            require_user_verification=True,
        )
    except Exception as e:
        logger.warning(f"WebAuthn registration verification fallita: {e}")
        return jsonify({'error': f'Verifica fallita: {str(e)}'}), 400

    # Protezione contro response=null: il secondo .get() fallirebbe su None
    response_obj = credential.get('response') or {}
    transports = ','.join(response_obj.get('transports') or [])
    db.webauthn_add_credential(
        credential_id=verification.credential_id,
        public_key=verification.credential_public_key,
        sign_count=verification.sign_count,
        nome=nome,
        transports=transports or None,
    )
    return jsonify({'success': True, 'nome': nome})


@app.route('/api/auth/webauthn/authenticate/begin', methods=['POST'])
def api_webauthn_auth_begin():
    """Step 1: genera options per il login con impronta (no auth richiesta)."""
    if not WEBAUTHN_AVAILABLE:
        return jsonify({'error': 'WebAuthn non disponibile'}), 500
    if not db.auth_is_configured():
        return jsonify({'error': 'Nessun utente configurato'}), 400

    creds = db.webauthn_get_credentials()
    if not creds:
        return jsonify({'error': 'Nessuna impronta registrata'}), 400

    allowed = [
        PublicKeyCredentialDescriptor(id=bytes(c['credential_id']))
        for c in creds
    ]

    options = generate_authentication_options(
        rp_id=WEBAUTHN_RP_ID,
        allow_credentials=allowed,
        user_verification=UserVerificationRequirement.REQUIRED,
    )
    session['webauthn_auth_challenge'] = bytes_to_base64url(options.challenge)
    return app.response_class(options_to_json(options), mimetype='application/json')


@app.route('/api/auth/webauthn/authenticate/complete', methods=['POST'])
def api_webauthn_auth_complete():
    """Step 2: verifica la risposta e autentica."""
    if not WEBAUTHN_AVAILABLE:
        return jsonify({'error': 'WebAuthn non disponibile'}), 500

    challenge_b64 = session.pop('webauthn_auth_challenge', None)
    if not challenge_b64:
        return jsonify({'error': 'Challenge non trovata'}), 400

    data = request.get_json() or {}
    credential = data.get('credential')
    if not credential:
        return jsonify({'error': 'Credential mancante'}), 400

    try:
        raw_id = base64url_to_bytes(credential['rawId'])
    except Exception:
        return jsonify({'error': 'rawId non valido'}), 400

    stored = db.webauthn_get_credential_by_id(raw_id)
    if not stored:
        return jsonify({'error': 'Credenziale non riconosciuta'}), 401

    try:
        verification = verify_authentication_response(
            credential=credential,
            expected_challenge=base64url_to_bytes(challenge_b64),
            expected_origin=WEBAUTHN_ORIGIN,
            expected_rp_id=WEBAUTHN_RP_ID,
            credential_public_key=bytes(stored['public_key']),
            credential_current_sign_count=stored['sign_count'],
            require_user_verification=True,
        )
    except Exception as e:
        logger.warning(f"WebAuthn auth verification fallita: {e}")
        return jsonify({'error': 'Verifica impronta fallita'}), 401

    db.webauthn_update_sign_count(raw_id, verification.new_sign_count)
    _login_user('webauthn')
    return jsonify({'success': True})


@app.route('/api/auth/webauthn/credentials', methods=['GET'])
@login_required
def api_webauthn_list_credentials():
    """Lista le credenziali registrate (solo metadata)."""
    creds = db.webauthn_get_credentials()
    return jsonify({
        'credentials': [
            {
                'id': c['id'],
                'nome': c['nome'],
                'data_registrazione': c['data_registrazione'],
                'ultimo_utilizzo': c['ultimo_utilizzo'],
            }
            for c in creds
        ]
    })


@app.route('/api/auth/webauthn/credentials/<int:cred_id>', methods=['DELETE'])
@login_required
def api_webauthn_delete_credential(cred_id):
    """Rimuove una credenziale registrata."""
    db.webauthn_delete_credential(cred_id)
    return jsonify({'success': True})


@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def api_auth_change_password():
    """Cambia la password dell'utente (richiede password corrente)."""
    data = request.get_json() or {}
    current = data.get('current_password') or ''
    nuova = data.get('new_password') or ''

    if len(nuova) < 8:
        return jsonify({'error': 'La nuova password deve essere almeno 8 caratteri'}), 400

    user = db.auth_get_user()
    if not user:
        # Scenario difensivo: record eliminato in un altro processo
        return jsonify({'error': 'Utente non trovato'}), 401
    if not check_password_hash(user['password_hash'], current):
        return jsonify({'error': 'Password corrente non valida'}), 401

    db.auth_update_credentials(password_hash=generate_password_hash(nuova))
    return jsonify({'success': True})


@app.route('/profilo')
@login_required
def profilo_page():
    """Pagina gestione profilo e impronte."""
    return render_template('profilo.html')




# Backup all'avvio
if config.BACKUP_ON_STARTUP:
    db.create_backup()
    logger.info("Backup automatico all'avvio completato")

# Costanti da config
MESI_NOME = config.MESI_NOME
MESI_SCOLASTICI = config.MESI_SCOLASTICI


# ==================== VALIDAZIONE ====================
# Le funzioni di validazione vivono in validators.py (condivise con i blueprint).
from validators import validate_string, validate_number  # noqa: E402


# ==================== ROUTES PAGINE ====================

@app.route('/')
def index():
    """Dashboard principale"""
    return render_template('index.html')


@app.route('/import')
def import_page():
    """Pagina import Excel"""
    return render_template('import.html')


@app.route('/rendicontazione')
def rendicontazione_page():
    """Pagina rendicontazione mensile"""
    return render_template('rendicontazione.html')


@app.route('/chiusura-mese')
def chiusura_mese_page():
    """Procedura guidata di chiusura del mese"""
    return render_template('chiusura_mese.html')


@app.route('/calendario')
def calendario_page():
    """Pagina gestione calendario scolastico"""
    return render_template('calendario.html')


@app.route('/report')
def report_page():
    """Pagina generazione report"""
    return render_template('report.html')


@app.route('/utenti')
def utenti_page():
    """Pagina gestione utenti"""
    return render_template('utenti.html')


@app.route('/dipendenti')
def dipendenti_page():
    """Anagrafica dipendenti (operatori OEPAC)"""
    return render_template('dipendenti.html')


@app.route('/dipendente/<int:dipendente_id>')
def dipendente_dettaglio_page(dipendente_id):
    """Scheda del singolo dipendente"""
    return render_template('dipendente_dettaglio.html', dipendente_id=dipendente_id)


@app.route('/turni')
def turni_page():
    """Planner turni settimanali degli operatori"""
    return render_template('turni.html')


@app.route('/sostituzioni')
def sostituzioni_page():
    """Assenze degli operatori e gestione delle sostituzioni"""
    return render_template('sostituzioni.html')


@app.route('/impostazioni')
def impostazioni_page():
    """Area Impostazioni: porta alla prima sezione (Commesse)."""
    return redirect(url_for('commesse_page'))


@app.route('/commesse')
def commesse_page():
    """Pagina gestione commesse"""
    return render_template('commesse.html')


@app.route('/statistiche')
def statistiche_page():
    """Pagina statistiche avanzate"""
    return render_template('statistiche.html')


# ==================== API ====================

@app.route('/api/import-excel/template')
def api_import_template():
    """Scarica template Excel per l'import"""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Utenti"

        # Header
        headers = ['Commessa', 'Albero Attività', 'Attività', 'Monte Ore Sett.']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Esempi
        ws.cell(row=2, column=1, value="CIG 123456")
        ws.cell(row=2, column=2, value="IC Esempio - Scuola Primaria")
        ws.cell(row=2, column=3, value="Rossi Mario")
        ws.cell(row=2, column=4, value=10)

        ws.cell(row=3, column=1, value="CIG 123456")
        ws.cell(row=3, column=2, value="IC Esempio - Scuola Secondaria")
        ws.cell(row=3, column=3, value="Bianchi Anna")
        ws.cell(row=3, column=4, value=15)

        # Larghezza colonne
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15

        # Salva in buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='template_import_utenti.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/import-excel/cronologia')
def api_import_cronologia():
    """Ottiene cronologia import"""
    try:
        with db.get_db_context() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT * FROM import_cronologia
                ORDER BY data DESC
                LIMIT 20
            ''')
            rows = cursor.fetchall()

            cronologia = []
            for r in rows:
                cronologia.append({
                    'id': r['id'],
                    'data': r['data'],
                    'filename': r['filename'],
                    'importati': r['importati'],
                    'aggiornati': r['aggiornati'],
                    'errori': r['errori'],
                    'utente': r['utente'] or 'Sistema'
                })

            return jsonify({'cronologia': cronologia})
    except Exception:
        # Se la tabella non esiste, la creiamo
        try:
            with db.get_db_context() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS import_cronologia (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        data TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        filename TEXT,
                        importati INTEGER DEFAULT 0,
                        aggiornati INTEGER DEFAULT 0,
                        errori INTEGER DEFAULT 0,
                        utente TEXT
                    )
                ''')
            return jsonify({'cronologia': []})
        except Exception:
            return jsonify({'cronologia': []})


@app.route('/api/import-excel/preview', methods=['POST'])
def api_preview_excel():
    """Anteprima import Excel - mostra cosa verra' importato senza modificare il DB"""
    if 'file' not in request.files:
        return jsonify({'error': 'Nessun file caricato'}), 400

    file = request.files['file']
    if file.filename == '' or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'File non valido'}), 400

    try:
        df = pd.read_excel(file)
        df.columns = [str(col).strip().lower() for col in df.columns]

        # Mappa colonne (stessa logica di import)
        col_map = {}
        for col in df.columns:
            col_lower = str(col).lower()
            if 'commessa' in col_lower:
                col_map['commessa'] = col
            elif 'albero' in col_lower:
                col_map['scuola'] = col
            elif 'scuola' in col_lower or 'ic' in col_lower or 'plesso' in col_lower:
                if 'scuola' not in col_map:
                    col_map['scuola'] = col
            elif col_lower == 'attività' or col_lower == 'attivita':
                col_map['utente'] = col
            elif 'nome' in col_lower and 'cognome' not in col_lower:
                if 'utente' not in col_map:
                    col_map['nome'] = col
            elif 'cognome' in col_lower:
                col_map['cognome'] = col
            elif 'utente' in col_lower or 'nominativo' in col_lower:
                col_map['utente'] = col
            elif 'monte' in col_lower or 'ore' in col_lower:
                col_map['monte_ore'] = col

        # Genera anteprima righe
        preview_rows = []
        errors = []

        for idx, row in df.iterrows():
            try:
                commessa = str(row.get(col_map.get('commessa', ''), '')).strip()
                scuola = str(row.get(col_map.get('scuola', ''), '')).strip()

                if 'utente' in col_map:
                    nome_completo = str(row[col_map['utente']]).strip()
                    parti = nome_completo.split()
                    # Formato Excel: "Cognome Nome" -> cognome = primo token, nome = resto
                    cognome = parti[0] if parti else ''
                    nome = ' '.join(parti[1:]) if len(parti) >= 2 else ''
                else:
                    nome = str(row.get(col_map.get('nome', ''), '')).strip()
                    cognome = str(row.get(col_map.get('cognome', ''), '')).strip()

                monte_ore_val = row.get(col_map.get('monte_ore', ''), None)
                if pd.isna(monte_ore_val) if monte_ore_val is not None else True:
                    continue
                monte_ore = float(monte_ore_val)

                if not nome or nome == 'nan':
                    continue
                if not scuola or scuola == 'nan':
                    continue

                # Verifica se l'utente esiste già
                stato = 'nuovo'
                try:
                    with db.get_db_context() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            SELECT u.id FROM utenti u
                            JOIN scuole s ON u.scuola_id = s.id
                            WHERE u.nome = ? AND u.cognome = ? AND s.nome LIKE ?
                        ''', (nome, cognome, f'%{scuola[:30]}%'))
                        if cursor.fetchone():
                            stato = 'esistente'
                except Exception as e:
                    logger.warning(f"Anteprima import, controllo esistenza utente fallito "
                                   f"(riga {idx + 2}, {nome} {cognome}): {e}")

                preview_rows.append({
                    'riga': idx + 2,
                    'commessa': commessa,
                    'scuola': scuola[:50],
                    'nome': nome,
                    'cognome': cognome,
                    'monte_ore': monte_ore,
                    'stato': stato
                })
            except Exception as e:
                errors.append(f"Riga {idx+2}: {str(e)}")

        return jsonify({
            'success': True,
            'colonne_trovate': list(df.columns),
            'colonne_mappate': col_map,
            'totale_righe': len(df),
            'righe_valide': len(preview_rows),
            'preview': preview_rows[:50],
            'errors': errors[:10]
        })

    except Exception as e:
        return jsonify({'error': f'Errore lettura file: {str(e)}'}), 500


@app.route('/api/import-excel', methods=['POST'])
def api_import_excel():
    """Importa dati da file Excel"""
    logger.info("Inizio import Excel")

    if 'file' not in request.files:
        return jsonify({'error': 'Nessun file caricato'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nessun file selezionato'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Formato file non valido. Usa .xlsx o .xls'}), 400

    # Rete di sicurezza: snapshot del DB prima di scrivere, per poter annullare
    # l'import ripristinando questo backup.
    backup_pre_import = db.create_backup()
    if backup_pre_import:
        logger.info(f"Backup pre-import creato: {backup_pre_import}")

    try:
        logger.info(f"Lettura file: {file.filename}")

        # Leggi Excel
        df = pd.read_excel(file)
        logger.info(f"Righe trovate: {len(df)}")
        logger.info(f"Colonne originali: {list(df.columns)}")

        # Normalizza nomi colonne (converti tutto a stringa prima)
        df.columns = [str(col).strip().lower() for col in df.columns]
        logger.info(f"Colonne normalizzate: {list(df.columns)}")

        # Trova le colonne necessarie
        col_map = {}
        for col in df.columns:
            col_lower = str(col).lower()
            if 'commessa' in col_lower:
                col_map['commessa'] = col
            elif 'albero' in col_lower:
                col_map['scuola'] = col
            elif 'scuola' in col_lower or 'ic' in col_lower or 'plesso' in col_lower:
                if 'scuola' not in col_map:
                    col_map['scuola'] = col
            elif col_lower == 'attività' or col_lower == 'attivita':
                col_map['utente'] = col
            elif 'nome' in col_lower and 'cognome' not in col_lower:
                if 'utente' not in col_map:
                    col_map['nome'] = col
            elif 'cognome' in col_lower:
                col_map['cognome'] = col
            elif 'utente' in col_lower or 'nominativo' in col_lower:
                col_map['utente'] = col
            elif 'monte' in col_lower or 'ore' in col_lower:
                col_map['monte_ore'] = col

        logger.info(f"Colonne mappate: {col_map}")

        # Verifica colonne necessarie
        required = ['commessa', 'monte_ore']
        if 'utente' not in col_map and ('nome' not in col_map or 'cognome' not in col_map):
            return jsonify({'error': f'Colonna nome utente non trovata. Colonne trovate: {list(df.columns)}'}), 400

        if 'scuola' not in col_map:
            return jsonify({'error': f'Colonna scuola non trovata. Colonne trovate: {list(df.columns)}'}), 400

        for req in required:
            if req not in col_map:
                return jsonify({'error': f'Colonna {req} non trovata. Colonne trovate: {list(df.columns)}'}), 400

        # Modalità import
        import_mode = request.form.get('mode', 'update')  # 'update' or 'skip'

        # Importa dati
        imported = 0
        updated = 0
        skipped = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                commessa = str(row[col_map['commessa']]).strip()
                scuola = str(row[col_map['scuola']]).strip()

                # Gestione nome
                if 'utente' in col_map:
                    nome_completo = str(row[col_map['utente']]).strip()
                    parti = nome_completo.split()
                    # Formato Excel: "Cognome Nome" -> cognome = primo token, nome = resto
                    if len(parti) >= 2:
                        cognome = parti[0]
                        nome = ' '.join(parti[1:])
                    else:
                        cognome = nome_completo
                        nome = ''
                else:
                    nome = str(row[col_map['nome']]).strip()
                    cognome = str(row[col_map.get('cognome', '')]).strip() if 'cognome' in col_map else ''

                monte_ore_val = row[col_map['monte_ore']]
                # Gestisci monte_ore che potrebbe essere stringa o numero
                if pd.isna(monte_ore_val):
                    continue
                monte_ore = float(monte_ore_val)

                # Salta righe vuote
                if not nome or nome == 'nan' or nome.lower() == 'nan':
                    continue
                if not scuola or scuola == 'nan' or scuola.lower() == 'nan':
                    continue

                # Crea/ottieni scuola
                scuola_id = db.get_or_create_scuola(commessa, scuola)
                if not scuola_id:
                    errors.append(f"Riga {idx+2}: Commessa '{commessa}' non valida")
                    continue

                # Verifica se l'utente esiste già
                with db.get_db_context() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT id FROM utenti
                        WHERE nome = ? AND cognome = ? AND scuola_id = ?
                    ''', (nome, cognome, scuola_id))
                    existing = cursor.fetchone()

                    if existing:
                        if import_mode == 'skip':
                            skipped += 1
                            continue
                        else:
                            # Aggiorna utente esistente
                            cursor.execute('''
                                UPDATE utenti SET monte_ore_settimanale = ?
                                WHERE id = ?
                            ''', (monte_ore, existing['id']))
                            updated += 1
                    else:
                        # Crea nuovo utente
                        db.get_or_create_utente(scuola_id, nome, cognome, monte_ore)
                        imported += 1

            except Exception as e:
                errors.append(f"Riga {idx+2}: {str(e)}")

        logger.info(f"Import completato: {imported} nuovi, {updated} aggiornati, {skipped} saltati")
        db.log_audit('import', 'utenti', dettagli=f'{imported} nuovi, {updated} aggiornati da {file.filename}')

        # Salva in cronologia
        try:
            with db.get_db_context() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS import_cronologia (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        data TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        filename TEXT,
                        importati INTEGER DEFAULT 0,
                        aggiornati INTEGER DEFAULT 0,
                        errori INTEGER DEFAULT 0,
                        utente TEXT
                    )
                ''')
                cursor.execute('''
                    INSERT INTO import_cronologia (filename, importati, aggiornati, errori)
                    VALUES (?, ?, ?, ?)
                ''', (file.filename, imported, updated, len(errors)))
        except Exception as e:
            logger.warning(f"Errore salvataggio cronologia: {e}")

        return jsonify({
            'success': True,
            'imported': imported,
            'updated': updated,
            'skipped': skipped,
            'errors': errors
        })

    except Exception as e:
        logger.error(f"Errore import: {str(e)}", exc_info=True)
        return jsonify({'error': f'Errore durante l\'import: {str(e)}'}), 500


# ==================== IMPORT RENDICONTAZIONE MENSILE ====================

def _analizza_rendicontazione(file):
    """Legge il file Excel e abbina le righe agli utenti in anagrafica.
    Ritorna la lista di fogli analizzati (vedi import_rendicontazione.analizza)."""
    fogli = import_rendicontazione.parse_workbook(file)
    utenti = db.get_all_utenti(page=None)
    return import_rendicontazione.analizza(fogli, utenti)


def _riepilogo_foglio(f, max_esempi=15):
    """Compatta un foglio analizzato per la risposta JSON (anteprima)."""
    def slim(voce):
        return {
            'nome': voce['nome_completo'],
            'scuola': (voce.get('scuola') or '')[:60],
            'commessa': voce.get('commessa', ''),
            'ore': voce.get('ore'),
            'pasti': voce.get('pasti'),
            'utente_nome': voce.get('utente_nome'),
        }
    return {
        'foglio': f['foglio'],
        'mese': f['mese'],
        'anno': f['anno'],
        'mese_nome': MESI_NOME.get(f['mese'], '') if f['mese'] else '',
        'riconosciuto': f['riconosciuto'],
        'periodo_valido': bool(f['mese'] and f['anno']),
        'n_match': len(f['match']),
        'n_non_trovati': len(f['non_trovati']),
        'n_ambigui': len(f['ambigui']),
        'n_senza_ore': len(f['senza_ore']),
        'anteprima_match': [slim(v) for v in f['match'][:max_esempi]],
        'esempi_non_trovati': [slim(v) for v in f['non_trovati'][:max_esempi]],
        'esempi_ambigui': [slim(v) for v in f['ambigui'][:max_esempi]],
    }


@app.route('/api/import-rendicontazione/preview', methods=['POST'])
def api_preview_rendicontazione():
    """Anteprima import rendicontazioni: mostra per ogni mese quanti utenti
    sono stati abbinati, senza modificare il database."""
    if 'file' not in request.files:
        return jsonify({'error': 'Nessun file caricato'}), 400
    file = request.files['file']
    if file.filename == '' or not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'File non valido. Usa .xlsx o .xls'}), 400

    try:
        analisi = _analizza_rendicontazione(file)
        fogli = [_riepilogo_foglio(f) for f in analisi]
        return jsonify({
            'success': True,
            'fogli': fogli,
            'totale_match': sum(f['n_match'] for f in fogli),
            'totale_non_trovati': sum(f['n_non_trovati'] for f in fogli),
            'totale_ambigui': sum(f['n_ambigui'] for f in fogli),
        })
    except Exception as e:
        logger.error(f"Errore anteprima rendicontazione: {e}", exc_info=True)
        return jsonify({'error': f'Errore lettura file: {str(e)}'}), 500


@app.route('/api/import-rendicontazione', methods=['POST'])
def api_import_rendicontazione():
    """Importa le rendicontazioni mensili. Scrive ore e pasti per gli utenti
    abbinati, un mese (foglio) alla volta in transazione."""
    if 'file' not in request.files:
        return jsonify({'error': 'Nessun file caricato'}), 400
    file = request.files['file']
    if file.filename == '' or not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'File non valido. Usa .xlsx o .xls'}), 400

    # Modalita': 'overwrite' sovrascrive sempre, 'solo_vuoti' tocca solo chi non ha ore
    modalita = (request.form.get('modalita') or 'solo_vuoti').strip()

    try:
        analisi = _analizza_rendicontazione(file)
    except Exception as e:
        logger.error(f"Errore lettura rendicontazione: {e}", exc_info=True)
        return jsonify({'error': 'Errore nella lettura del file Excel. Controlla che sia un file valido.'}), 500

    # Rete di sicurezza: snapshot del DB prima di scrivere le ore (annullabile via restore)
    backup_pre_import = db.create_backup()
    if backup_pre_import:
        logger.info(f"Backup pre-import rendicontazione creato: {backup_pre_import}")

    dettaglio = []
    tot_scritti = tot_saltati_pieni = tot_non_trovati = tot_ambigui = 0

    for f in analisi:
        if not (f['mese'] and f['anno']):
            dettaglio.append({'foglio': f['foglio'], 'errore': 'Mese/anno non riconosciuto dal foglio'})
            continue

        anno, mese = f['anno'], f['mese']

        # In modalita' "solo vuoti" salta gli utenti che hanno gia' ore nel mese
        gia_con_ore = set()
        if modalita != 'overwrite':
            for d in db.get_rendicontazione_completa(anno, mese):
                if (d.get('ore_lavorate_60') or 0) > 0:
                    gia_con_ore.add(d['utente_id'])

        updates = []
        saltati_pieni = 0
        for voce in f['match']:
            if voce['utente_id'] in gia_con_ore:
                saltati_pieni += 1
                continue
            updates.append({
                'utente_id': voce['utente_id'],
                'ore_lavorate': voce['ore'],
                'pasti': voce['pasti'],
            })

        try:
            scritti = db.update_rendicontazione_batch(anno, mese, updates)
        except Exception as e:
            logger.error(f"Errore scrittura rendicontazione {mese}/{anno}: {e}", exc_info=True)
            dettaglio.append({'foglio': f['foglio'], 'mese': mese, 'anno': anno,
                              'errore': str(e)})
            continue

        tot_scritti += scritti
        tot_saltati_pieni += saltati_pieni
        tot_non_trovati += len(f['non_trovati'])
        tot_ambigui += len(f['ambigui'])
        dettaglio.append({
            'foglio': f['foglio'], 'mese': mese, 'anno': anno,
            'mese_nome': MESI_NOME.get(mese, ''),
            'scritti': scritti,
            'saltati_gia_pieni': saltati_pieni,
            'non_trovati': len(f['non_trovati']),
            'ambigui': len(f['ambigui']),
        })

    db.log_audit('import_rendicontazione', 'rendicontazione',
                 dettagli=f"{tot_scritti} rendicontazioni importate da {file.filename}")
    try:
        with db.get_db_context() as conn:
            conn.execute('''
                CREATE TABLE IF NOT EXISTS import_cronologia (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    data TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    filename TEXT, importati INTEGER DEFAULT 0,
                    aggiornati INTEGER DEFAULT 0, errori INTEGER DEFAULT 0, utente TEXT
                )''')
            conn.execute('''INSERT INTO import_cronologia (filename, importati, aggiornati, errori, utente)
                            VALUES (?, ?, 0, ?, ?)''',
                         (f"[Rendicontazione] {file.filename}", tot_scritti,
                          tot_non_trovati + tot_ambigui, session.get('username', 'Sistema')))
    except Exception as e:
        logger.warning(f"Cronologia import rendicontazione non salvata: {e}")

    return jsonify({
        'success': True,
        'modalita': modalita,
        'totale_scritti': tot_scritti,
        'totale_saltati_gia_pieni': tot_saltati_pieni,
        'totale_non_trovati': tot_non_trovati,
        'totale_ambigui': tot_ambigui,
        'dettaglio': dettaglio,
    })


# ==================== API DIPENDENTI ====================

@app.route('/api/dipendenti', methods=['GET'])
def api_get_dipendenti():
    """Elenco dipendenti, con ricerca/filtro opzionali e conteggio assistiti."""
    include_inactive = request.args.get('tutti') == '1'
    search = request.args.get('q')
    commessa_id = request.args.get('commessa_id', type=int)
    dipendenti = db.get_all_dipendenti(include_inactive=include_inactive, search=search, commessa_id=commessa_id)
    conteggi = db.count_assegnazioni_bulk()
    for d in dipendenti:
        d.update(conteggi.get(d['id'], {'assistiti': 0, 'ore_assegnate': 0}))
    return jsonify(dipendenti)


@app.route('/api/dipendenti', methods=['POST'])
def api_create_dipendente():
    """Crea un dipendente."""
    data = request.json or {}
    try:
        dip_id = db.create_dipendente(data)
        db.log_audit('create', 'dipendente', dip_id, f"{data.get('cognome','')} {data.get('nome','')}")
        return jsonify({'success': True, 'id': dip_id})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/dipendenti/<int:dipendente_id>', methods=['GET'])
def api_get_dipendente(dipendente_id):
    d = db.get_dipendente(dipendente_id)
    if not d:
        return jsonify({'error': 'Dipendente non trovato'}), 404
    d['assistiti'] = db.get_assistiti_dipendente(dipendente_id)
    # solo le ore (il conteggio 'assistiti' lo ricava il frontend dalla lista)
    d['ore_assegnate'] = db.count_assegnazioni_dipendente(dipendente_id)['ore_assegnate']
    return jsonify(d)


@app.route('/api/dipendenti/<int:dipendente_id>', methods=['PUT'])
def api_update_dipendente(dipendente_id):
    data = request.json or {}
    try:
        db.update_dipendente(dipendente_id, data)
        db.log_audit('update', 'dipendente', dipendente_id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/dipendenti/<int:dipendente_id>', methods=['DELETE'])
def api_delete_dipendente(dipendente_id):
    db.delete_dipendente(dipendente_id)
    db.log_audit('delete', 'dipendente', dipendente_id)
    return jsonify({'success': True})


@app.route('/api/import-dipendenti/preview', methods=['POST'])
def api_preview_dipendenti():
    """Anteprima import anagrafica dipendenti: quanti nuovi/da aggiornare."""
    if 'file' not in request.files:
        return jsonify({'error': 'Nessun file caricato'}), 400
    file = request.files['file']
    if file.filename == '' or not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'File non valido. Usa .xlsx o .xls'}), 400
    try:
        records = import_dipendenti.parse_workbook(file.read())
    except Exception as e:
        logger.error(f"Errore anteprima dipendenti: {e}", exc_info=True)
        return jsonify({'error': f'Errore lettura file: {str(e)}'}), 500

    # Abbina agli esistenti (per CF o nome+cognome) per contare nuovi/da aggiornare
    esistenti = db.get_all_dipendenti(include_inactive=True)
    per_cf = {import_dipendenti.chiave_cf(d['codice_fiscale']): d for d in esistenti if d.get('codice_fiscale')}
    per_nome = {import_dipendenti.chiave_nome(d['nome'], d['cognome']): d for d in esistenti}

    nuovi = aggiornati = 0
    colonne_extra = set()
    esempi = []
    for r in records:
        cf = import_dipendenti.chiave_cf(r.get('codice_fiscale'))
        if cf:
            match = per_cf.get(cf)
        else:
            match = per_nome.get(import_dipendenti.chiave_nome(r.get('nome', ''), r.get('cognome', '')))
        if match:
            aggiornati += 1
        else:
            nuovi += 1
        colonne_extra.update((r.get('extra') or {}).keys())
        if len(esempi) < 12:
            esempi.append({
                'nome': f"{r.get('cognome','')} {r.get('nome','')}".strip(),
                'cf': r.get('codice_fiscale') or '',
                'qualifica': r.get('qualifica') or '',
                'ore': r.get('ore_contrattuali_settimanali'),
                'stato': 'aggiorna' if match else 'nuovo',
            })

    return jsonify({
        'success': True,
        'totale': len(records),
        'nuovi': nuovi,
        'aggiornati': aggiornati,
        'colonne_extra': sorted(colonne_extra),
        'esempi': esempi,
    })


@app.route('/api/import-dipendenti', methods=['POST'])
def api_import_dipendenti():
    """Importa l'anagrafica dipendenti (crea i nuovi, aggiorna gli esistenti)."""
    if 'file' not in request.files:
        return jsonify({'error': 'Nessun file caricato'}), 400
    file = request.files['file']
    if file.filename == '' or not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'File non valido. Usa .xlsx o .xls'}), 400
    try:
        records = import_dipendenti.parse_workbook(file.read())
        res = db.importa_dipendenti(records)
    except Exception as e:
        logger.error(f"Errore import dipendenti: {e}", exc_info=True)
        return jsonify({'error': f'Errore: {str(e)}'}), 500

    db.log_audit('import_dipendenti', 'dipendente',
                 dettagli=f"{res['creati']} creati, {res['aggiornati']} aggiornati da {file.filename}")
    try:
        with db.get_db_context() as conn:
            conn.execute('''CREATE TABLE IF NOT EXISTS import_cronologia (
                id INTEGER PRIMARY KEY AUTOINCREMENT, data TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                filename TEXT, importati INTEGER DEFAULT 0, aggiornati INTEGER DEFAULT 0,
                errori INTEGER DEFAULT 0, utente TEXT)''')
            conn.execute('''INSERT INTO import_cronologia (filename, importati, aggiornati, errori, utente)
                            VALUES (?, ?, ?, ?, ?)''',
                         (f"[Dipendenti] {file.filename}", res['creati'], res['aggiornati'],
                          len(res['errori']), session.get('username', 'Sistema')))
    except Exception as e:
        logger.warning(f"Cronologia import dipendenti non salvata: {e}")

    return jsonify({'success': True, **res, 'errori_count': len(res['errori']),
                    'errori': res['errori'][:10]})


# ==================== API ASSEGNAZIONI (utente <-> operatore) ====================

@app.route('/api/utente/<int:utente_id>/assegnazioni', methods=['GET'])
def api_get_assegnazioni_utente(utente_id):
    """Operatori assegnati a un assistito + bilancio ore."""
    return jsonify({
        'assegnazioni': db.get_assegnazioni_utente(utente_id),
        'bilancio': db.get_bilancio_assegnazioni_utente(utente_id),
    })


@app.route('/api/utente/<int:utente_id>/assegnazioni', methods=['POST'])
def api_create_assegnazione(utente_id):
    data = request.json or {}
    dipendente_id = data.get('dipendente_id')
    if not dipendente_id:
        return jsonify({'error': 'Operatore obbligatorio'}), 400
    try:
        ore = float(data.get('ore_settimanali') or 0)
    except (ValueError, TypeError):
        return jsonify({'error': 'Ore non valide'}), 400
    try:
        aid = db.create_assegnazione(utente_id, dipendente_id, ore,
                                     data.get('valido_da'), data.get('valido_a'), data.get('note'))
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    db.log_audit('create', 'assegnazione', aid, f"utente {utente_id} -> dip {dipendente_id} ({ore}h)")
    return jsonify({'success': True, 'id': aid, 'bilancio': db.get_bilancio_assegnazioni_utente(utente_id)})


@app.route('/api/assegnazioni/<int:assegnazione_id>', methods=['PUT'])
def api_update_assegnazione(assegnazione_id):
    data = request.json or {}
    ore = data.get('ore_settimanali')
    try:
        db.update_assegnazione(
            assegnazione_id,
            ore_settimanali=float(ore) if ore is not None else None,
            valido_da=data.get('valido_da'),
            valido_a=data.get('valido_a'),
            note=data.get('note'),
        )
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    return jsonify({'success': True})


@app.route('/api/assegnazioni/<int:assegnazione_id>', methods=['DELETE'])
def api_delete_assegnazione(assegnazione_id):
    db.delete_assegnazione(assegnazione_id)
    return jsonify({'success': True})


# ==================== API TURNI ====================

@app.route('/api/dipendenti/<int:dipendente_id>/turni', methods=['GET'])
def api_get_turni_dipendente(dipendente_id):
    return jsonify({
        'turni': db.get_turni_dipendente(dipendente_id),
        'ore_pianificate': db.ore_settimanali_pianificate(dipendente_id),
        'giorni': db.GIORNI_NOMI,
    })


@app.route('/api/dipendenti/<int:dipendente_id>/turni', methods=['POST'])
def api_create_turno(dipendente_id):
    d = request.json or {}
    if d.get('giorno') is None or not d.get('ora_inizio') or not d.get('ora_fine'):
        return jsonify({'error': 'Giorno e orari sono obbligatori'}), 400
    try:
        tid = db.create_turno(dipendente_id, d['giorno'], d['ora_inizio'], d['ora_fine'],
                              d.get('scuola_id'), d.get('utente_id'),
                              d.get('valido_da'), d.get('valido_a'), d.get('note'))
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    return jsonify({'success': True, 'id': tid})


@app.route('/api/turni/<int:turno_id>', methods=['PUT'])
def api_update_turno(turno_id):
    try:
        db.update_turno(turno_id, **(request.json or {}))
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    return jsonify({'success': True})


@app.route('/api/turni/<int:turno_id>', methods=['DELETE'])
def api_delete_turno(turno_id):
    db.delete_turno(turno_id)
    return jsonify({'success': True})


@app.route('/api/turni/giorno')
def api_turni_giorno():
    """Vista giornaliera: tutti i turni di un giorno della settimana."""
    giorno = request.args.get('giorno', type=int)
    data = request.args.get('data')
    scuola_id = request.args.get('scuola_id', type=int)
    if giorno is None:
        return jsonify({'error': 'Giorno obbligatorio'}), 400
    return jsonify(db.get_turni_giorno(giorno, data=data, scuola_id=scuola_id))


# ==================== API ASSENZE DIPENDENTI ====================

@app.route('/api/dipendenti/<int:dipendente_id>/assenze', methods=['GET'])
def api_get_assenze_dip(dipendente_id):
    return jsonify(db.get_assenze_dipendente(dipendente_id))


@app.route('/api/dipendenti/<int:dipendente_id>/assenze', methods=['POST'])
def api_create_assenza_dip(dipendente_id):
    d = request.json or {}
    if not d.get('data_inizio'):
        return jsonify({'error': 'Data inizio obbligatoria'}), 400
    try:
        aid = db.create_assenza_dipendente(dipendente_id, d['data_inizio'], d.get('data_fine'),
                                           d.get('tipo'), d.get('motivazione'), d.get('note'))
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    # Genera automaticamente le sostituzioni da coprire per i turni nel periodo
    create = db.crea_sostituzioni_per_assenza(aid)
    db.log_audit('create', 'assenza_dipendente', aid, f"dip {dipendente_id}, {create} turni da coprire")
    return jsonify({'success': True, 'id': aid, 'turni_da_coprire': create})


@app.route('/api/assenze-dipendenti/<int:assenza_id>', methods=['DELETE'])
def api_delete_assenza_dip(assenza_id):
    db.delete_assenza_dipendente(assenza_id)
    return jsonify({'success': True})


# ==================== API SOSTITUZIONI ====================

@app.route('/api/sostituzioni', methods=['GET'])
def api_get_sostituzioni():
    return jsonify(db.get_sostituzioni(
        data_inizio=request.args.get('da'),
        data_fine=request.args.get('a'),
        solo_da_coprire=request.args.get('da_coprire') == '1',
    ))


@app.route('/api/sostituzioni/<int:sostituzione_id>/candidati', methods=['GET'])
def api_candidati_sostituzione(sostituzione_id):
    s = db.get_sostituzione(sostituzione_id)
    if not s:
        return jsonify({'error': 'Sostituzione non trovata'}), 404
    candidati = db.suggerisci_sostituti(s.get('scuola_id'), s['giorno'], s['ora_inizio'],
                                        s['ora_fine'], s['data'], escludi_id=s['assente_id'])
    return jsonify({'candidati': candidati})


@app.route('/api/sostituzioni/<int:sostituzione_id>/assegna', methods=['POST'])
def api_assegna_sostituto(sostituzione_id):
    d = request.json or {}
    if not d.get('sostituto_id'):
        return jsonify({'error': 'Sostituto obbligatorio'}), 400
    try:
        db.assegna_sostituto(sostituzione_id, d['sostituto_id'])
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    return jsonify({'success': True})


@app.route('/api/sostituzioni/<int:sostituzione_id>/annulla', methods=['POST'])
def api_annulla_sostituto(sostituzione_id):
    db.annulla_sostituto(sostituzione_id)
    return jsonify({'success': True})


@app.route('/api/utenti', methods=['GET'])
def api_get_utenti():
    """Ottiene lista utenti con paginazione opzionale"""
    commessa = request.args.get('commessa')
    scuola_id = request.args.get('scuola_id')
    page = request.args.get('page', type=int)
    limit = request.args.get('limit', default=50, type=int)

    # Limita il numero massimo di elementi per pagina
    limit = min(limit, 200)

    utenti = db.get_all_utenti(commessa, scuola_id, page=page, limit=limit)

    # Se paginato, restituisci anche i metadati
    if page is not None:
        total = db.count_utenti(commessa, scuola_id)
        return jsonify({
            'data': [dict(u) for u in utenti],
            'pagination': {
                'page': page,
                'limit': limit,
                'total': total,
                'pages': (total + limit - 1) // limit
            }
        })

    return jsonify([dict(u) for u in utenti])


@app.route('/api/scuole', methods=['GET'])
def api_get_scuole():
    """Ottiene lista scuole"""
    commessa = request.args.get('commessa')
    scuole = db.get_all_scuole(commessa)
    return jsonify(scuole)


@app.route('/api/utenti', methods=['POST'])
def api_create_utente():
    """Crea un nuovo utente manualmente"""
    data = request.json or {}

    # Validazione
    errors = []

    commessa, err = validate_string(data.get('commessa'), 'Commessa', config.MAX_COMMESSA_LENGTH)
    if err: errors.append(err)

    scuola_nome, err = validate_string(data.get('scuola'), 'Scuola', config.MAX_SCUOLA_LENGTH)
    if err: errors.append(err)

    nome, err = validate_string(data.get('nome'), 'Nome', config.MAX_NOME_LENGTH)
    if err: errors.append(err)

    cognome, err = validate_string(data.get('cognome'), 'Cognome', config.MAX_COGNOME_LENGTH, required=False)
    if err: errors.append(err)

    monte_ore, err = validate_number(data.get('monte_ore'), 'Monte ore', 0, config.MAX_ORE_SETTIMANALI)
    if err: errors.append(err)

    if errors:
        return jsonify({'error': '; '.join(errors)}), 400

    try:
        scuola_id = db.get_or_create_scuola(commessa, scuola_nome)
        if not scuola_id:
            return jsonify({'error': f'Commessa "{commessa}" non valida'}), 400

        utente_id = db.get_or_create_utente(scuola_id, nome, cognome, monte_ore)
        db.log_audit('creazione', 'utente', utente_id, f'{nome} {cognome}')
        logger.info(f"Utente creato: {nome} {cognome} (ID: {utente_id})")

        return jsonify({'success': True, 'utente_id': utente_id})
    except Exception as e:
        logger.error(f"Errore creazione utente: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/utenti/<int:utente_id>', methods=['PUT'])
def api_update_utente(utente_id):
    """Aggiorna un utente esistente"""
    data = request.json or {}

    try:
        with db.get_db_context() as conn:
            cursor = conn.cursor()

            cursor.execute("SELECT * FROM utenti WHERE id = ?", (utente_id,))
            utente = cursor.fetchone()
            if not utente:
                return jsonify({'error': 'Utente non trovato'}), 404

            dati_precedenti = dict(utente)

            # Salva per undo
            push_undo('update_utente', {
                'id': utente_id,
                'dati_precedenti': dati_precedenti
            })

            updates = []
            params = []
            errors = []

            if 'nome' in data and data['nome']:
                nome, err = validate_string(data['nome'], 'Nome', config.MAX_NOME_LENGTH)
                if err:
                    errors.append(err)
                else:
                    updates.append("nome = ?")
                    params.append(nome)

            if 'cognome' in data:
                cognome, err = validate_string(data['cognome'], 'Cognome', config.MAX_COGNOME_LENGTH, required=False)
                if err:
                    errors.append(err)
                else:
                    updates.append("cognome = ?")
                    params.append(cognome)

            if 'monte_ore' in data and data['monte_ore'] is not None:
                monte_ore, err = validate_number(data['monte_ore'], 'Monte ore', 0, config.MAX_ORE_SETTIMANALI)
                if err:
                    errors.append(err)
                else:
                    updates.append("monte_ore_settimanale = ?")
                    params.append(monte_ore)

            if 'lista_attesa' in data:
                updates.append("lista_attesa = ?")
                params.append(data['lista_attesa'] if data['lista_attesa'] else None)

            # Gestione periodo di validità (data_inizio, data_fine)
            if 'data_inizio' in data:
                # Formato atteso: 'YYYY-MM' o null/vuoto per rimuovere
                data_inizio = data['data_inizio'].strip() if data['data_inizio'] else None
                if data_inizio and not re.match(r'^\d{4}-\d{2}$', data_inizio):
                    errors.append("data_inizio deve essere nel formato YYYY-MM")
                else:
                    updates.append("data_inizio = ?")
                    params.append(data_inizio)

            if 'data_fine' in data:
                # Formato atteso: 'YYYY-MM' o null/vuoto per rimuovere
                data_fine = data['data_fine'].strip() if data['data_fine'] else None
                if data_fine and not re.match(r'^\d{4}-\d{2}$', data_fine):
                    errors.append("data_fine deve essere nel formato YYYY-MM")
                else:
                    updates.append("data_fine = ?")
                    params.append(data_fine)

            if errors:
                return jsonify({'error': '; '.join(errors)}), 400

            if 'nome' in data or 'cognome' in data:
                # Usa "or" per gestire sia valori None espliciti nel JSON che valori mancanti
                nome = (data.get('nome') or utente['nome'] or '').strip()
                cognome = (data.get('cognome') or utente['cognome'] or '').strip()
                nome_puntato = db.punteggia_nome(nome, cognome)
                updates.append("nome_puntato = ?")
                params.append(nome_puntato)

            if not updates:
                return jsonify({'success': True, 'message': 'Nessuna modifica da applicare'})

            params.append(utente_id)
            cursor.execute(f"UPDATE utenti SET {', '.join(updates)} WHERE id = ?", params)

        db.log_audit('modifica', 'utente', utente_id, dettagli='Aggiornamento utente', dati_precedenti=dati_precedenti, dati_nuovi=data)
        logger.info(f"Utente aggiornato: ID {utente_id}")

        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Errore aggiornamento utente {utente_id}: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/utenti/<int:utente_id>', methods=['DELETE'])
def api_delete_utente(utente_id):
    """Elimina un utente"""
    try:
        with db.get_db_context() as conn:
            cursor = conn.cursor()
            # Snapshot per undo (utente + rendicontazioni + note/documenti/assenze)
            snapshot = db.raccogli_snapshot_utente(cursor, utente_id)
            # Cancellazione completa (FK enforcement attivo: vanno rimossi
            # anche i record correlati senza CASCADE, nella stessa transazione)
            db.elimina_utente_completo(cursor, utente_id)

        # push_undo SOLO dopo il commit riuscito: una delete fallita non deve
        # lasciare un'azione fantasma nello stack (il cui undo fallirebbe)
        if snapshot:
            push_undo('delete_utente', snapshot)

        u = snapshot['utente'] if snapshot else None
        nome_utente = f"{u['nome']} {u['cognome']}" if u else f"ID {utente_id}"
        db.log_audit('eliminazione', 'utente', utente_id, f'Eliminato utente {nome_utente}')
        logger.info(f"Utente eliminato: {nome_utente} (ID: {utente_id})")

        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Errore eliminazione utente {utente_id}: {e}", exc_info=True)
        return jsonify({'error': "Errore nell'eliminazione dell'utente"}), 500


@app.route('/api/utenti/bulk', methods=['PUT'])
def api_bulk_update_utenti():
    """Aggiornamento bulk di utenti (monte ore o lista attesa)"""
    data = request.json or {}
    updates = data.get('updates', [])

    if not updates:
        return jsonify({'error': 'Nessun aggiornamento specificato'}), 400

    if len(updates) > 200:
        return jsonify({'error': 'Massimo 200 aggiornamenti per richiesta'}), 400

    successi = 0
    errori = []

    for upd in updates:
        uid = upd.get('id')
        if not uid:
            continue
        try:
            set_parts = []
            params = []

            if 'monte_ore' in upd and upd['monte_ore'] is not None:
                ore, err = validate_number(upd['monte_ore'], 'Monte ore', 0, config.MAX_ORE_SETTIMANALI)
                if err:
                    errori.append(f'ID {uid}: {err}')
                    continue
                set_parts.append('monte_ore_settimanale = ?')
                params.append(ore)

            if 'lista_attesa' in upd:
                set_parts.append('lista_attesa = ?')
                params.append(upd['lista_attesa'] if upd['lista_attesa'] else None)

            if set_parts:
                with db.get_db_context() as conn:
                    cursor = conn.cursor()
                    params.append(uid)
                    cursor.execute(f"UPDATE utenti SET {', '.join(set_parts)} WHERE id = ?", params)
                successi += 1
        except Exception as e:
            errori.append(f'ID {uid}: {str(e)}')

    db.log_audit('bulk_modifica', 'utenti', dettagli=f'{successi} utenti aggiornati in bulk')
    logger.info(f"Bulk update: {successi} successi, {len(errori)} errori")

    return jsonify({
        'success': True,
        'aggiornati': successi,
        'errori': errori
    })


@app.route('/api/utenti/bulk', methods=['DELETE'])
def api_bulk_delete_utenti():
    """Eliminazione bulk di utenti"""
    data = request.json or {}
    ids = data.get('ids', [])

    if not ids:
        return jsonify({'error': 'Nessun utente specificato'}), 400

    eliminati = 0
    for uid in ids:
        try:
            with db.get_db_context() as conn:
                cursor = conn.cursor()
                snapshot = db.raccogli_snapshot_utente(cursor, uid)
                db.elimina_utente_completo(cursor, uid)
            # push_undo solo a transazione riuscita (niente azioni fantasma)
            if snapshot:
                push_undo('delete_utente', snapshot)
            eliminati += 1
        except Exception as e:
            logger.error(f"Errore bulk delete utente {uid}: {e}")

    db.log_audit('bulk_eliminazione', 'utenti', dettagli=f'{eliminati} utenti eliminati in bulk')
    return jsonify({'success': True, 'eliminati': eliminati})


@app.route('/api/utenti/export-csv')
def api_export_utenti_csv():
    """Esporta lista utenti in formato CSV"""
    commessa = request.args.get('commessa')

    with db.get_db_context() as conn:
        cursor = conn.cursor()

        query = '''
            SELECT u.id, u.nome, u.cognome, u.monte_ore_settimanale, u.lista_attesa,
                   u.data_inizio, u.data_fine, u.attivo,
                   s.nome_completo as scuola, c.nome as commessa
            FROM utenti u
            LEFT JOIN scuole s ON u.scuola_id = s.id
            LEFT JOIN commesse c ON s.commessa_id = c.id
            WHERE 1=1
        '''
        params = []
        if commessa:
            query += ' AND c.nome = ?'
            params.append(commessa)
        query += ' ORDER BY u.cognome, u.nome'

        cursor.execute(query, params)
        utenti = cursor.fetchall()

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['nome', 'cognome', 'scuola', 'monte_ore', 'commessa', 'lista_attesa', 'data_inizio', 'data_fine', 'attivo'])

    for u in utenti:
        writer.writerow([
            u['nome'], u['cognome'], u['scuola'], u['monte_ore_settimanale'],
            u['commessa'], u['lista_attesa'] or '', u['data_inizio'] or '', u['data_fine'] or '',
            'si' if u['attivo'] else 'no'
        ])

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = 'attachment; filename=utenti_export.csv'
    return response


@app.route('/api/utenti/import-csv', methods=['POST'])
def api_import_utenti_csv():
    """Importa utenti da file CSV"""
    if 'file' not in request.files:
        return jsonify({'error': 'Nessun file caricato'}), 400

    file = request.files['file']
    mode = request.form.get('mode', 'add')  # 'add' o 'update'

    try:
        content = file.read().decode('utf-8-sig')  # utf-8-sig per gestire BOM
        reader = csv.DictReader(StringIO(content))

        aggiunti = 0
        aggiornati = 0
        errori = []

        with db.get_db_context() as conn:
            cursor = conn.cursor()

            for i, row in enumerate(reader, start=2):
                try:
                    nome = row.get('nome', '').strip()
                    cognome = row.get('cognome', '').strip()
                    scuola = row.get('scuola', '').strip()
                    monte_ore = float(row.get('monte_ore', 0) or 0)
                    commessa = row.get('commessa', '').strip()

                    if not nome:
                        errori.append(f"Riga {i}: nome mancante")
                        continue

                    # Trova o crea scuola
                    cursor.execute('''
                        SELECT s.id FROM scuole s
                        JOIN commesse c ON s.commessa_id = c.id
                        WHERE s.nome_completo = ? AND c.nome = ?
                    ''', (scuola, commessa))
                    scuola_row = cursor.fetchone()

                    if not scuola_row:
                        # Trova commessa
                        cursor.execute("SELECT id FROM commesse WHERE nome = ?", (commessa,))
                        comm = cursor.fetchone()
                        if not comm:
                            errori.append(f"Riga {i}: commessa '{commessa}' non trovata")
                            continue
                        # Crea scuola
                        cursor.execute("INSERT INTO scuole (nome_completo, commessa_id) VALUES (?, ?)", (scuola, comm['id']))
                        scuola_id = cursor.lastrowid
                    else:
                        scuola_id = scuola_row['id']

                    # Verifica se utente esiste
                    cursor.execute("SELECT id FROM utenti WHERE nome = ? AND cognome = ? AND scuola_id = ?", (nome, cognome, scuola_id))
                    existing = cursor.fetchone()

                    if existing and mode == 'update':
                        cursor.execute("UPDATE utenti SET monte_ore_settimanale = ? WHERE id = ?", (monte_ore, existing['id']))
                        aggiornati += 1
                    elif not existing:
                        nome_puntato = db.punteggia_nome(nome, cognome)
                        cursor.execute('''
                            INSERT INTO utenti (nome, cognome, nome_puntato, monte_ore_settimanale, scuola_id, attivo, data_inserimento)
                            VALUES (?, ?, ?, ?, ?, 1, ?)
                        ''', (nome, cognome, nome_puntato, monte_ore, scuola_id, datetime.now().isoformat()))
                        aggiunti += 1

                except Exception as e:
                    errori.append(f"Riga {i}: {str(e)}")

        return jsonify({
            'success': True,
            'aggiunti': aggiunti,
            'aggiornati': aggiornati,
            'errori': errori[:10]  # Max 10 errori
        })

    except Exception as e:
        return jsonify({'error': f'Errore elaborazione CSV: {str(e)}'}), 400


@app.route('/api/utenti/<int:utente_id>/duplica', methods=['POST'])
def api_duplica_utente(utente_id):
    """Duplica un utente"""
    with db.get_db_context() as conn:
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM utenti WHERE id = ?", (utente_id,))
        utente = cursor.fetchone()

        if not utente:
            return jsonify({'error': 'Utente non trovato'}), 404

        nuovo_nome = f"{utente['nome']} (copia)"
        nome_puntato = db.punteggia_nome(nuovo_nome, utente['cognome'])

        cursor.execute('''
            INSERT INTO utenti (nome, cognome, nome_puntato, monte_ore_settimanale, scuola_id, lista_attesa, attivo, data_inserimento)
            VALUES (?, ?, ?, ?, ?, ?, 1, ?)
        ''', (nuovo_nome, utente['cognome'], nome_puntato, utente['monte_ore_settimanale'], utente['scuola_id'], utente['lista_attesa'], datetime.now().isoformat()))

        nuovo_id = cursor.lastrowid

    db.log_audit('duplica', 'utente', nuovo_id, dettagli=f'Duplicato da utente {utente_id}')
    return jsonify({'success': True, 'nuovo_id': nuovo_id})


@app.route('/api/utenti/<int:utente_id>/cambio-scuola', methods=['POST'])
def api_cambio_scuola(utente_id):
    """Cambia la scuola di un utente"""
    data = request.json or {}
    nuova_scuola_id = data.get('scuola_id')

    if not nuova_scuola_id:
        return jsonify({'error': 'scuola_id richiesto'}), 400

    with db.get_db_context() as conn:
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM utenti WHERE id = ?", (utente_id,))
        utente = cursor.fetchone()
        if not utente:
            return jsonify({'error': 'Utente non trovato'}), 404

        cursor.execute("SELECT id FROM scuole WHERE id = ?", (nuova_scuola_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Scuola non trovata'}), 404

        vecchia_scuola = utente['scuola_id']
        cursor.execute("UPDATE utenti SET scuola_id = ? WHERE id = ?", (nuova_scuola_id, utente_id))

    db.log_audit('cambio_scuola', 'utente', utente_id, dettagli=f'Da scuola {vecchia_scuola} a {nuova_scuola_id}')
    return jsonify({'success': True})


@app.route('/api/utenti/<int:utente_id>/storico-monte-ore')
def api_storico_monte_ore(utente_id):
    """Ottiene lo storico delle modifiche al monte ore"""
    with db.get_db_context() as conn:
        cursor = conn.cursor()

        # Info utente
        cursor.execute("SELECT nome, cognome, monte_ore_settimanale FROM utenti WHERE id = ?", (utente_id,))
        utente = cursor.fetchone()
        if not utente:
            return jsonify({'error': 'Utente non trovato'}), 404

        # Cerca nei log di audit le modifiche al monte ore
        cursor.execute('''
            SELECT timestamp, dati_precedenti, dati_nuovi, dettagli
            FROM audit_log
            WHERE entita = 'utente' AND entita_id = ?
            AND (dati_nuovi LIKE '%monte_ore%' OR dettagli LIKE '%monte%')
            ORDER BY timestamp DESC
            LIMIT 20
        ''', (utente_id,))

        storico = []
        for row in cursor.fetchall():
            try:
                import json
                prec = json.loads(row['dati_precedenti']) if row['dati_precedenti'] else {}
                nuov = json.loads(row['dati_nuovi']) if row['dati_nuovi'] else {}
                storico.append({
                    'timestamp': row['timestamp'],
                    'monte_ore_precedente': prec.get('monte_ore_settimanale') or prec.get('monte_ore'),
                    'monte_ore_nuovo': nuov.get('monte_ore_settimanale') or nuov.get('monte_ore'),
                    'dettagli': row['dettagli']
                })
            except (json.JSONDecodeError, KeyError, TypeError):
                pass

        utente_info = {
            'nome': utente['nome'],
            'cognome': utente['cognome'],
            'monte_ore_attuale': utente['monte_ore_settimanale']
        }

    return jsonify({
        'utente': utente_info,
        'storico': storico
    })


# ==================== VARIAZIONI MONTE ORE ====================

@app.route('/api/utenti/<int:utente_id>/variazioni-monte-ore')
@login_required
def api_get_variazioni_monte_ore(utente_id):
    """Lista variazioni monte ore di un utente"""
    variazioni = db.get_variazioni_monte_ore(utente_id)
    return jsonify({'variazioni': variazioni})


@app.route('/api/utenti/<int:utente_id>/variazioni-monte-ore', methods=['POST'])
@login_required
def api_add_variazione_monte_ore(utente_id):
    """Aggiunge una variazione monte ore"""
    data = request.json or {}

    monte_ore, err = validate_number(data.get('monte_ore'), 'Monte ore', 0, config.MAX_ORE_SETTIMANALI)
    if err:
        return jsonify({'error': err}), 400

    mese_inizio = (data.get('mese_inizio') or '').strip()
    if not mese_inizio or len(mese_inizio) != 7:
        return jsonify({'error': 'Mese inizio obbligatorio (formato YYYY-MM)'}), 400

    nota = (data.get('nota') or '').strip()[:500] or None

    var_id = db.add_variazione_monte_ore(utente_id, monte_ore, mese_inizio, nota)
    db.log_audit('variazione_monte_ore', 'utente', utente_id,
                 dettagli=f'Aggiunta variazione: {monte_ore}h da {mese_inizio}')
    return jsonify({'success': True, 'id': var_id})


@app.route('/api/variazioni-monte-ore/<int:variazione_id>', methods=['PUT'])
@login_required
def api_update_variazione_monte_ore(variazione_id):
    """Modifica una variazione monte ore"""
    data = request.json or {}

    kwargs = {}
    if 'monte_ore' in data and data['monte_ore'] is not None:
        monte_ore, err = validate_number(data['monte_ore'], 'Monte ore', 0, config.MAX_ORE_SETTIMANALI)
        if err:
            return jsonify({'error': err}), 400
        kwargs['monte_ore'] = monte_ore

    if 'mese_inizio' in data:
        mese_inizio = (data['mese_inizio'] or '').strip()
        if not mese_inizio or len(mese_inizio) != 7:
            return jsonify({'error': 'Mese inizio non valido'}), 400
        kwargs['mese_inizio'] = mese_inizio

    if 'nota' in data:
        kwargs['nota'] = (data['nota'] or '').strip()[:500] or None

    if not kwargs:
        return jsonify({'error': 'Nessun campo da aggiornare'}), 400

    ok = db.update_variazione_monte_ore(variazione_id, **kwargs)
    if not ok:
        return jsonify({'error': 'Variazione non trovata'}), 404
    return jsonify({'success': True})


@app.route('/api/variazioni-monte-ore/<int:variazione_id>', methods=['DELETE'])
@login_required
def api_delete_variazione_monte_ore(variazione_id):
    """Elimina una variazione monte ore"""
    ok = db.delete_variazione_monte_ore(variazione_id)
    if not ok:
        return jsonify({'error': 'Variazione non trovata'}), 404
    return jsonify({'success': True})


@app.route('/api/reset', methods=['POST'])
def api_reset_data():
    """Reset completo dei dati - RICHIEDE conferma esplicita"""
    data = request.json or {}
    reset_type = data.get('type', 'all')
    confirm_text = data.get('confirm', '')

    # Richiedi conferma esplicita: l'utente deve digitare "CONFERMA"
    if confirm_text != 'CONFERMA':
        return jsonify({'error': 'Per procedere, devi confermare digitando "CONFERMA"'}), 400

    valid_types = ['all', 'rendicontazioni', 'utenti']
    if reset_type not in valid_types:
        return jsonify({'error': 'Tipo reset non valido'}), 400

    try:
        # Backup automatico prima del reset
        backup_name = db.create_backup()
        logger.warning(f"RESET richiesto (tipo: {reset_type}). Backup creato: {backup_name}")

        with db.get_db_context() as conn:
            cursor = conn.cursor()

            # Ordine FK-safe: con foreign_keys=ON i figli senza CASCADE
            # (documenti/note/assenze) vanno rimossi prima degli utenti, e i
            # turni scollegati da utenti/scuole prima di cancellarli.
            if reset_type == 'all':
                cursor.execute("DELETE FROM rendicontazione")
                cursor.execute("DELETE FROM documenti_utente")
                cursor.execute("DELETE FROM note_utente")
                cursor.execute("DELETE FROM assenze")
                cursor.execute("UPDATE turni SET utente_id = NULL, scuola_id = NULL")
                cursor.execute("DELETE FROM utenti")
                cursor.execute("DELETE FROM scuole")
                message = "Tutti i dati sono stati cancellati"
            elif reset_type == 'rendicontazioni':
                cursor.execute("DELETE FROM rendicontazione")
                message = "Tutte le rendicontazioni sono state cancellate"
            elif reset_type == 'utenti':
                cursor.execute("DELETE FROM rendicontazione")
                cursor.execute("DELETE FROM documenti_utente")
                cursor.execute("DELETE FROM note_utente")
                cursor.execute("DELETE FROM assenze")
                cursor.execute("UPDATE turni SET utente_id = NULL")
                cursor.execute("DELETE FROM utenti")
                message = "Utenti e rendicontazioni cancellati"

        db.log_audit('reset', 'sistema', dettagli=f'Reset {reset_type}: {message}. Backup: {backup_name}')
        logger.warning(f"Reset completato: {message}")

        return jsonify({'success': True, 'message': message, 'backup': backup_name})
    except Exception as e:
        logger.error(f"Errore reset: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/commesse', methods=['GET'])
def api_get_commesse():
    """Ottiene lista commesse"""
    commesse = db.get_all_commesse()
    return jsonify(commesse)


@app.route('/api/commesse', methods=['POST'])
def api_create_commessa():
    """Crea una nuova commessa"""
    data = request.json or {}
    errors = []

    nome, err = validate_string(data.get('nome'), 'Nome commessa', config.MAX_COMMESSA_LENGTH)
    if err: errors.append(err)

    descrizione, err = validate_string(data.get('descrizione', ''), 'Descrizione', config.MAX_DESCRIZIONE_LENGTH, required=False)
    if err: errors.append(err)

    colore = data.get('colore', '#6366f1')
    if colore and not re.match(r'^#[0-9a-fA-F]{6}$', colore):
        errors.append('Colore non valido (formato: #RRGGBB)')

    if errors:
        return jsonify({'error': '; '.join(errors)}), 400

    commessa_id = db.create_commessa(nome, descrizione, colore)
    if not commessa_id:
        return jsonify({'error': 'Commessa gia\' esistente'}), 400

    db.log_audit('creazione', 'commessa', commessa_id, f'Nuova commessa: {nome}')
    logger.info(f"Commessa creata: {nome} (ID: {commessa_id})")

    return jsonify({'success': True, 'id': commessa_id})


@app.route('/api/commesse/<int:commessa_id>', methods=['PUT'])
def api_update_commessa(commessa_id):
    """Aggiorna una commessa"""
    data = request.json or {}
    errors = []

    nome = data.get('nome')
    if nome is not None:
        nome, err = validate_string(nome, 'Nome commessa', config.MAX_COMMESSA_LENGTH)
        if err: errors.append(err)

    descrizione = data.get('descrizione')
    if descrizione is not None:
        descrizione, err = validate_string(descrizione, 'Descrizione', config.MAX_DESCRIZIONE_LENGTH, required=False)
        if err: errors.append(err)

    colore = data.get('colore')
    if colore and not re.match(r'^#[0-9a-fA-F]{6}$', colore):
        errors.append('Colore non valido (formato: #RRGGBB)')

    if errors:
        return jsonify({'error': '; '.join(errors)}), 400

    db.update_commessa(
        commessa_id,
        nome=nome,
        descrizione=descrizione,
        colore=colore,
        attiva=data.get('attiva')
    )
    return jsonify({'success': True})


@app.route('/api/commesse/<int:commessa_id>', methods=['DELETE'])
def api_delete_commessa(commessa_id):
    """Elimina una commessa (soft delete)"""
    db.delete_commessa(commessa_id)
    return jsonify({'success': True})


@app.route('/api/stats/advanced')
def api_stats_advanced():
    """Statistiche avanzate per dashboard"""
    anno = request.args.get('anno', type=int)
    mese = request.args.get('mese', type=int)
    commessa = request.args.get('commessa')
    stats = db.get_statistiche_avanzate(anno, mese, commessa)
    return jsonify(stats)


@app.route('/api/stats/utenti-meno-ore/<int:anno>/<int:mese>')
def api_utenti_meno_ore(anno, mese):
    """Ottiene i 10 utenti con meno ore erogate rispetto alle previste"""
    utenti = db.get_utenti_meno_ore(anno, mese, limit=10)
    return jsonify(utenti)


@app.route('/api/stats/ore-confronto/<anno_scolastico>')
def api_ore_confronto(anno_scolastico):
    """Ottiene il confronto ore erogate vs previste per l'anno scolastico"""
    commessa = request.args.get('commessa')
    dati = db.get_ore_erogate_vs_previste(anno_scolastico, commessa)
    return jsonify(dati)


@app.route('/api/search')
def api_search():
    """Ricerca globale"""
    query = request.args.get('q', '').strip().lower()
    if len(query) < 2:
        return jsonify({'utenti': [], 'scuole': [], 'pages': []})

    # Pagine statiche
    pages = []
    pages_data = [
        {'title': 'Dashboard', 'url': '/', 'keywords': 'home dashboard principale panoramica riepilogo'},
        {'title': 'Import Excel', 'url': '/import', 'keywords': 'import excel carica dati upload file anteprima'},
        {'title': 'Rendicontazione', 'url': '/rendicontazione', 'keywords': 'ore rendicontazione inserisci mensile lavorate pasti'},
        {'title': 'Utenti', 'url': '/utenti', 'keywords': 'utenti gestione elenco operatori bulk modifica'},
        {'title': 'Commesse', 'url': '/commesse', 'keywords': 'commesse progetti gestione contratti lotti'},
        {'title': 'Calendario', 'url': '/calendario', 'keywords': 'calendario giorni lavorativi anno scolastico'},
        {'title': 'Report', 'url': '/report', 'keywords': 'report esporta excel word pdf stampa rendiconto'},
        {'title': 'Statistiche', 'url': '/statistiche', 'keywords': 'statistiche grafici analisi trend confronto'},
        {'title': 'Backup e Ripristino', 'url': '/statistiche', 'keywords': 'backup ripristino salvataggio dati restore'},
        {'title': 'Audit Trail', 'url': '/statistiche', 'keywords': 'audit trail log modifiche cronologia storico'},
    ]
    for p in pages_data:
        if query in p['keywords'] or query in p['title'].lower():
            pages.append({'title': p['title'], 'url': p['url']})

    # Cerca utenti
    with db.get_db_context() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT u.id, u.nome, u.cognome, c.nome as commessa, s.nome_completo as scuola
            FROM utenti u
            JOIN scuole s ON u.scuola_id = s.id
            JOIN commesse c ON s.commessa_id = c.id
            WHERE u.attivo = 1 AND (
                LOWER(u.nome) LIKE ? OR
                LOWER(u.cognome) LIKE ? OR
                LOWER(u.nome || ' ' || u.cognome) LIKE ?
            )
            LIMIT 10
        ''', (f'%{query}%', f'%{query}%', f'%{query}%'))
        utenti = [dict(r) for r in cursor.fetchall()]

        # Cerca scuole
        cursor.execute('''
            SELECT s.id, s.nome_completo, c.nome as commessa
            FROM scuole s
            JOIN commesse c ON s.commessa_id = c.id
            WHERE c.attiva = 1 AND LOWER(s.nome_completo) LIKE ?
            LIMIT 10
        ''', (f'%{query}%',))
        scuole = [dict(r) for r in cursor.fetchall()]

        # Cerca commesse
        cursor.execute('''
            SELECT c.id, c.nome, c.descrizione, c.colore,
                (SELECT COUNT(*) FROM scuole s WHERE s.commessa_id = c.id) as num_scuole
            FROM commesse c
            WHERE c.attiva = 1 AND (LOWER(c.nome) LIKE ? OR LOWER(c.descrizione) LIKE ?)
            LIMIT 5
        ''', (f'%{query}%', f'%{query}%'))
        commesse = [dict(r) for r in cursor.fetchall()]

    return jsonify({
        'pages': pages[:5],
        'utenti': utenti,
        'scuole': scuole,
        'commesse': commesse
    })


@app.route('/api/rendicontazione/<int:anno>/<int:mese>', methods=['GET'])
def api_get_rendicontazione(anno, mese):
    """Ottiene rendicontazione mensile"""
    commessa = request.args.get('commessa')
    dati = db.get_rendicontazione_completa(anno, mese, commessa)
    totali_scuola = db.get_totali_per_scuola(anno, mese, commessa)

    # Calcola totali generali. Imponibile/IVA/totale sono ricalcolati UNA volta
    # sulla somma delle ore (stesso metodo di export e totali per scuola): sommare
    # i per-riga gia' arrotondati divergerebbe di centesimi dai documenti ufficiali.
    ore_totali_100 = sum(d['ore_lavorate_100'] or 0 for d in dati)
    imponibile_tot, iva_tot, totale_tot = config.calcola_fatturazione(ore_totali_100)
    totale_generale = {
        'ore_lavorate_60': sum(d['ore_lavorate_60'] or 0 for d in dati),
        'ore_lavorate_100': ore_totali_100,
        'imponibile_100': imponibile_tot,
        'iva_100': iva_tot,
        'totale_100': totale_tot,
        'pasti': sum(d['pasti'] or 0 for d in dati),
        'credito_debito': round(sum(d['credito_debito'] or 0 for d in dati), 2),
        'num_utenti': len(dati)
    }

    return jsonify({
        'dati': dati,
        'totali_scuola': totali_scuola,
        'totale_generale': totale_generale,
        'mese_nome': MESI_NOME.get(mese, ''),
        'anno': anno
    })


@app.route('/api/rendicontazione/<int:anno>/<int:mese>', methods=['POST'])
def api_update_rendicontazione(anno, mese):
    """Aggiorna rendicontazione per un utente"""
    data = request.json
    utente_id = data.get('utente_id')

    if not utente_id:
        return jsonify({'error': 'utente_id richiesto'}), 400

    # Assicurati che esista la rendicontazione
    db.get_or_create_rendicontazione(utente_id, anno, mese)

    # Aggiorna
    db.update_rendicontazione(
        utente_id, anno, mese,
        ore_lavorate=data.get('ore_lavorate_60'),
        pasti=data.get('pasti'),
        note=data.get('note')
    )

    return jsonify({'success': True})


@app.route('/api/rendicontazione/<int:anno>/<int:mese>/batch', methods=['POST'])
def api_batch_update_rendicontazione(anno, mese):
    """Aggiorna rendicontazione per più utenti in una singola transazione"""
    data = request.json
    updates = data.get('updates', [])

    try:
        aggiornati = db.update_rendicontazione_batch(anno, mese, [
            {
                'utente_id': u.get('utente_id'),
                'ore_lavorate': u.get('ore_lavorate_60'),
                'pasti': u.get('pasti'),
                'note': u.get('note')
            }
            for u in updates
        ])
    except Exception as e:
        return jsonify({'success': False, 'error': f'Salvataggio annullato: {e}'}), 500

    return jsonify({'success': True, 'aggiornati': aggiornati})


@app.route('/api/rendicontazione/<int:anno>/<int:mese>/copia-precedente', methods=['POST'])
def api_copia_mese_precedente(anno, mese):
    """Copia ore dal mese precedente per utenti selezionati o tutti"""
    data = request.json
    utente_ids = data.get('utente_ids', [])  # Se vuoto, copia per tutti
    solo_vuoti = data.get('solo_vuoti', True)  # Copia solo se utente non ha ore

    # Calcola mese precedente
    if mese == 1:
        mese_prec = 12
        anno_prec = anno - 1
    else:
        mese_prec = mese - 1
        anno_prec = anno

    # Ottieni dati mese precedente
    dati_prec = db.get_rendicontazione_completa(anno_prec, mese_prec)
    if not dati_prec:
        return jsonify({'error': 'Nessun dato nel mese precedente', 'copiati': 0}), 404

    # Ottieni dati mese corrente per verificare quali utenti hanno già ore
    dati_corrente = db.get_rendicontazione_completa(anno, mese)
    utenti_con_ore = {d['utente_id'] for d in dati_corrente if (d.get('ore_lavorate_60') or 0) > 0}

    updates = []
    for d in dati_prec:
        uid = d['utente_id']

        # Filtra per utente_ids se specificato
        if utente_ids and uid not in utente_ids:
            continue

        # Salta se utente ha già ore e solo_vuoti è True
        if solo_vuoti and uid in utenti_con_ore:
            continue

        ore_prec = d.get('ore_lavorate_60', 0)
        if ore_prec and ore_prec > 0:
            updates.append({'utente_id': uid, 'ore_lavorate': ore_prec})

    try:
        copiati = db.update_rendicontazione_batch(anno, mese, updates)
    except Exception as e:
        return jsonify({'success': False, 'error': f'Copia annullata: {e}'}), 500

    return jsonify({
        'success': True,
        'copiati': copiati,
        'mese_origine': f"{MESI_NOME.get(mese_prec, '')} {anno_prec}"
    })


@app.route('/api/rendicontazione/<int:anno>/<int:mese>/compila-media', methods=['POST'])
def api_compila_con_media(anno, mese):
    """Compila le ore con la media prevista per utenti senza ore"""
    data = request.json or {}
    utente_ids = data.get('utente_ids', [])  # Se vuoto, compila per tutti senza ore

    dati = db.get_rendicontazione_completa(anno, mese)
    updates = []

    for d in dati:
        uid = d['utente_id']

        # Filtra per utente_ids se specificato
        if utente_ids and uid not in utente_ids:
            continue

        # Solo utenti senza ore
        if (d.get('ore_lavorate_60') or 0) > 0:
            continue

        # Usa la media con assenza come valore predefinito
        media = d.get('media_con_assenza_60', 0)
        if media and media > 0:
            updates.append({'utente_id': uid, 'ore_lavorate': round(media, 2)})

    try:
        compilati = db.update_rendicontazione_batch(anno, mese, updates)
    except Exception as e:
        return jsonify({'success': False, 'error': f'Compilazione annullata: {e}'}), 500

    return jsonify({
        'success': True,
        'compilati': compilati
    })


@app.route('/api/utente/<int:utente_id>/storico')
def api_storico_utente(utente_id):
    """Storico ore degli ultimi N mesi a ritroso, con media prevista e
    credito/debito calcolati per ogni riga (tiene conto delle variazioni
    monte ore). Complementare a `/storico-ore` (vedi nota in api_get_storico_utente)."""
    mesi_da_mostrare = request.args.get('mesi', config.STORICO_MESI_DEFAULT, type=int)

    with db.get_db_context() as conn:
        cursor = conn.cursor()

        # Info utente
        cursor.execute('''
            SELECT u.*, s.nome_completo as scuola, c.nome as commessa
            FROM utenti u
            LEFT JOIN scuole s ON u.scuola_id = s.id
            LEFT JOIN commesse c ON s.commessa_id = c.id
            WHERE u.id = ?
        ''', (utente_id,))
        utente = cursor.fetchone()

        if not utente:
            return jsonify({'error': 'Utente non trovato'}), 404

        # Storico rendicontazione: i giorni effettivi dipendono dal tipo di scuola
        scuola_utente = utente['scuola'] or ''
        is_infanzia = 'INFANZIA' in scuola_utente.upper()
        cursor.execute('''
            SELECT r.*,
                CASE WHEN ? = 1 THEN cal.giorni_lavorativi
                     ELSE COALESCE(cal.giorni_lavorativi_altri, cal.giorni_lavorativi)
                END as giorni_lavorativi
            FROM rendicontazione r
            LEFT JOIN calendario_scolastico cal ON r.mese = cal.mese AND r.anno = cal.anno
            WHERE r.utente_id = ?
            ORDER BY r.anno DESC, r.mese DESC
            LIMIT ?
        ''', (1 if is_infanzia else 0, utente_id, mesi_da_mostrare))

        storico = []
        monte_ore_base = utente['monte_ore_settimanale'] or 0
        variazioni_utente = db.get_variazioni_monte_ore(utente_id)

        def _get_monte_ore_mese(anno_r, mese_r):
            periodo = f"{anno_r:04d}-{mese_r:02d}"
            mo = monte_ore_base
            for v in variazioni_utente:
                if v['mese_inizio'] <= periodo:
                    mo = v['monte_ore']
            return mo

        for r in cursor.fetchall():
            monte_ore = _get_monte_ore_mese(r['anno'], r['mese'])
            # Regola unica condivisa con la vista mensile (calendario -> default)
            giorni = db.risolvi_giorni_lavorativi(r['giorni_lavorativi'])
            media_60, media_assenza = db.calcola_media_prevista(monte_ore, giorni)
            ore = r['ore_lavorate_60'] or 0
            credito_debito = media_assenza - ore

            storico.append({
                'anno': r['anno'],
                'mese': r['mese'],
                'mese_nome': MESI_NOME.get(r['mese'], ''),
                'ore_lavorate_60': round(ore, 2),
                'ore_lavorate_100': round(ore, 2),
                'media_prevista': round(media_assenza, 2),
                'credito_debito': round(credito_debito, 2),
                'pasti': r['pasti'] or 0,
                'note': r['note']
            })

        utente_info = {
            'id': utente['id'],
            'nome': utente['nome'],
            'cognome': utente['cognome'],
            'monte_ore_settimanale': utente['monte_ore_settimanale'],
            'scuola': utente['scuola'],
            'commessa': utente['commessa']
        }

    return jsonify({
        'utente': utente_info,
        'storico': storico
    })


@app.route('/api/rendicontazione/<int:anno>/<int:mese>/confronto-precedente')
def api_confronto_mese_precedente(anno, mese):
    """Ottiene le differenze di ore rispetto al mese precedente"""
    commessa = request.args.get('commessa')

    # Calcola mese precedente
    if mese == 1:
        mese_prec, anno_prec = 12, anno - 1
    else:
        mese_prec, anno_prec = mese - 1, anno

    dati_corrente = db.get_rendicontazione_completa(anno, mese, commessa)
    dati_prec = db.get_rendicontazione_completa(anno_prec, mese_prec, commessa)

    # Crea mappa ore mese precedente
    ore_prec_map = {d['utente_id']: d.get('ore_lavorate_60', 0) for d in dati_prec}

    differenze = {}
    for d in dati_corrente:
        uid = d['utente_id']
        ore_corr = d.get('ore_lavorate_60', 0) or 0
        ore_prec = ore_prec_map.get(uid, 0) or 0

        if ore_prec > 0:
            diff = ore_corr - ore_prec
            diff_perc = round((diff / ore_prec) * 100, 1) if ore_prec else 0
        else:
            diff = ore_corr
            diff_perc = 100 if ore_corr > 0 else 0

        differenze[uid] = {
            'ore_precedente': round(ore_prec, 2),
            'ore_corrente': round(ore_corr, 2),
            'differenza': round(diff, 2),
            'differenza_perc': diff_perc
        }

    return jsonify({
        'differenze': differenze,
        'mese_precedente': {
            'anno': anno_prec,
            'mese': mese_prec,
            'mese_nome': MESI_NOME.get(mese_prec, '')
        }
    })


@app.route('/api/alerts', methods=['GET'])
def api_get_alerts():
    """Ottiene alert automatici per la dashboard"""
    anno = request.args.get('anno', type=int)
    mese = request.args.get('mese', type=int)

    if not anno or not mese:
        # Default: mese corrente
        from datetime import datetime
        now = datetime.now()
        anno = now.year
        mese = now.month

    alerts = []

    # 1. Utenti senza ore nel mese corrente
    dati = db.get_rendicontazione_completa(anno, mese)
    senza_ore = [d for d in dati if not d.get('ore_lavorate_60') or d['ore_lavorate_60'] == 0]
    if senza_ore:
        alerts.append({
            'type': 'warning',
            'title': f'{len(senza_ore)} utenti senza ore registrate',
            'message': f'Nel mese di {MESI_NOME.get(mese, "")} {anno}',
            'action': '/rendicontazione',
            'action_label': 'Vai a Rendicontazione',
            'count': len(senza_ore)
        })

    # 2. Utenti con ore molto sotto la media (< 50%)
    sotto_media = [d for d in dati
                   if (d.get('ore_lavorate_60') or 0) > 0
                   and (d.get('media_con_assenza_60') or 0) > 0
                   and d['ore_lavorate_60'] < d['media_con_assenza_60'] * 0.5]
    if sotto_media:
        alerts.append({
            'type': 'danger',
            'title': f'{len(sotto_media)} utenti con ore < 50% della media',
            'message': 'Ore erogate significativamente sotto la media prevista',
            'action': '/rendicontazione',
            'action_label': 'Verifica',
            'count': len(sotto_media),
            'utenti': [f"{d['nome']} {d['cognome']}" for d in sotto_media[:5]]
        })

    # 3. Utenti con ore anomale (> 150% della media)
    sopra_media = [d for d in dati
                   if (d.get('ore_lavorate_60') or 0) > 0
                   and (d.get('media_con_assenza_60') or 0) > 0
                   and d['ore_lavorate_60'] > d['media_con_assenza_60'] * 1.5]
    if sopra_media:
        alerts.append({
            'type': 'info',
            'title': f'{len(sopra_media)} utenti con ore > 150% della media',
            'message': 'Verificare se le ore extra sono corrette',
            'action': '/rendicontazione',
            'action_label': 'Verifica',
            'count': len(sopra_media),
            'utenti': [f"{d['nome']} {d['cognome']}" for d in sopra_media[:5]]
        })

    # 4. Verifica calendario - giorni lavorativi non impostati
    anno_scolastico = f"{anno}-{anno+1}" if mese >= 9 else f"{anno-1}-{anno}"
    with db.get_db_context() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT COUNT(*) as count FROM calendario_scolastico
            WHERE anno_scolastico = ? AND giorni_lavorativi > 0
        ''', (anno_scolastico,))
        cal_count = cursor.fetchone()['count']

    if cal_count < 10:  # Meno di 10 mesi configurati
        alerts.append({
            'type': 'warning',
            'title': 'Calendario scolastico incompleto',
            'message': f'Solo {cal_count}/10 mesi configurati per {anno_scolastico}',
            'action': '/calendario',
            'action_label': 'Configura Calendario',
            'count': 10 - cal_count
        })

    # 5. Percentuale completamento mese
    if dati:
        completati = len([d for d in dati if (d.get('ore_lavorate_60') or 0) > 0])
        perc = round((completati / len(dati)) * 100)
        if perc < 100:
            alerts.append({
                'type': 'info',
                'title': f'Completamento mese: {perc}%',
                'message': f'{completati}/{len(dati)} utenti rendicontati',
                'action': '/rendicontazione',
                'action_label': 'Completa',
                'progress': perc
            })

    # 6. Documenti in scadenza nei prossimi 30 giorni
    try:
        docs = db.get_documenti_in_scadenza(giorni=30)
        if docs:
            alerts.append({
                'type': 'warning',
                'title': f'{len(docs)} documenti in scadenza',
                'message': 'Documenti utente in scadenza nei prossimi 30 giorni',
                'count': len(docs),
                'utenti': [f"{d['nome']} {d['cognome']}" for d in docs[:5]]
            })
    except Exception as e:
        logger.warning(f"Alert documenti in scadenza non disponibile: {e}")

    # 7. Budget ore quasi esaurito
    try:
        critici = db.get_utenti_budget_critico(anno_scolastico, 80)
        if critici:
            alerts.append({
                'type': 'danger',
                'title': f"{len(critici)} utenti oltre l'80% del budget ore",
                'message': f'Budget annuale quasi esaurito ({anno_scolastico})',
                'count': len(critici),
                'utenti': [f"{u['nome']} {u['cognome']}" for u in critici[:5]]
            })
    except Exception as e:
        logger.warning(f"Alert budget critici non disponibile: {e}")

    return jsonify({
        'alerts': alerts,
        'anno': anno,
        'mese': mese,
        'mese_nome': MESI_NOME.get(mese, ''),
        'total_alerts': len(alerts)
    })


@app.route('/api/calendario', methods=['GET'])
def api_get_calendario():
    """Ottiene calendario scolastico"""
    anno_scolastico = request.args.get('anno_scolastico', '2025-2026')

    with db.get_db_context() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT * FROM calendario_scolastico
            WHERE anno_scolastico = ?
            ORDER BY anno, mese
        ''', (anno_scolastico,))
        result = cursor.fetchall()

        calendario = []
        for r in result:
            calendario.append({
                'mese': r['mese'],
                'anno': r['anno'],
                'giorni_lavorativi': r['giorni_lavorativi'],
                'giorni_lavorativi_altri': r['giorni_lavorativi_altri'],
                'mese_nome': MESI_NOME.get(r['mese'], '')
            })

    return jsonify({
        'anno_scolastico': anno_scolastico,
        'calendario': calendario
    })


@app.route('/api/calendario', methods=['POST'])
def api_update_calendario():
    """Aggiorna giorni lavorativi (default/infanzia + opzionale altri)"""
    data = request.json
    # giorni_lavorativi_altri e' opzionale: se non passato o None -> non distinzione
    altri = data.get('giorni_lavorativi_altri')
    if altri == '' or altri == 0:
        altri = None
    db.set_calendario(
        data['anno_scolastico'],
        data['mese'],
        data['anno'],
        data['giorni_lavorativi'],
        altri
    )
    return jsonify({'success': True})


@app.route('/api/calendario/<anno_scolastico>/copia-precedente', methods=['POST'])
def api_copia_calendario_precedente(anno_scolastico):
    """Copia i giorni lavorativi dall'anno precedente"""
    try:
        anno_inizio, anno_fine = map(int, anno_scolastico.split('-'))
        anno_precedente = f"{anno_inizio - 1}-{anno_inizio}"

        with db.get_db_context() as conn:
            cursor = conn.cursor()

            # Verifica che l'anno precedente esista
            cursor.execute('''
                SELECT COUNT(*) as count FROM calendario_scolastico
                WHERE anno_scolastico = ?
            ''', (anno_precedente,))
            if cursor.fetchone()['count'] == 0:
                return jsonify({'error': f'Anno {anno_precedente} non trovato'}), 404

            # Ottieni dati dell'anno precedente
            cursor.execute('''
                SELECT mese, giorni_lavorativi, giorni_lavorativi_altri FROM calendario_scolastico
                WHERE anno_scolastico = ?
            ''', (anno_precedente,))
            dati_precedenti = cursor.fetchall()

            copiati = 0
            for row in dati_precedenti:
                mese = row['mese']
                giorni = row['giorni_lavorativi']
                giorni_altri = row['giorni_lavorativi_altri']

                # Calcola l'anno corretto per questo mese
                if mese >= 9:  # Settembre-Dicembre
                    anno = anno_inizio
                else:  # Gennaio-Giugno
                    anno = anno_fine

                # Inserisci o aggiorna
                cursor.execute('''
                    INSERT OR REPLACE INTO calendario_scolastico
                    (anno_scolastico, mese, anno, giorni_lavorativi, giorni_lavorativi_altri)
                    VALUES (?, ?, ?, ?, ?)
                ''', (anno_scolastico, mese, anno, giorni, giorni_altri))
                copiati += 1

        return jsonify({
            'success': True,
            'copiati': copiati,
            'da_anno': anno_precedente
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/calendario/<anno_scolastico>/calcola-auto', methods=['POST'])
def api_calcola_calendario_auto(anno_scolastico):
    """Calcola automaticamente i giorni lavorativi (lun-ven, escluse festivita').
    Per giugno calcola anche giorni_lavorativi_altri fino all'8 giugno (fine scuola
    primaria/secondaria), mentre il campo principale copre fino al 30 giugno (infanzia).
    Per gli altri mesi usa date tipiche italiane (scuole aperte)."""
    try:
        from datetime import date, timedelta
        anno_inizio, anno_fine = map(int, anno_scolastico.split('-'))

        # Festivita' fisse italiane (giorno, mese) escluse dai giorni lavorativi
        FESTIVITA_FISSE = {
            (1, 1),   # Capodanno
            (6, 1),   # Epifania
            (25, 4),  # Liberazione
            (1, 5),   # Festa del Lavoro
            (2, 6),   # Festa della Repubblica
            (29, 6),  # SS. Pietro e Paolo (patrono di Roma)
            (1, 11),  # Ognissanti
            (8, 12),  # Immacolata
            (25, 12), # Natale
            (26, 12), # Santo Stefano
        }

        def calcola_pasqua(anno):
            """Data della domenica di Pasqua (algoritmo di Gauss/Butcher)."""
            a = anno % 19
            b = anno // 100
            c = anno % 100
            d = b // 4
            e = b % 4
            f = (b + 8) // 25
            g = (b - f + 1) // 3
            h = (19 * a + b - d - g + 15) % 30
            i = c // 4
            k = c % 4
            l = (32 + 2 * e + 2 * i - h - k) % 7
            m = (a + 11 * h + 22 * l) // 451
            mese = (h + l - 7 * m + 114) // 31
            giorno = ((h + l - 7 * m + 114) % 31) + 1
            return date(anno, mese, giorno)

        # Vacanze pasquali (calendario Lazio): dal giovedi' santo
        # al martedi' dopo Pasqua, inclusi
        pasqua = calcola_pasqua(anno_fine)
        vacanze_pasquali = set()
        d = pasqua - timedelta(days=3)
        while d <= pasqua + timedelta(days=2):
            vacanze_pasquali.add(d)
            d += timedelta(days=1)

        # Finestre di calendario scolastico (inizio, fine) per ogni mese, inclusive.
        # Valori tipici per regione Lazio; il numero risultante e' comunque editabile manualmente.
        # Per giugno si usano DUE finestre: infanzia fino al 30, altri fino all'8.
        FINESTRE = {
            9:  (date(anno_inizio, 9, 15), date(anno_inizio, 9, 30)),   # inizio scuola ~15/9
            10: (date(anno_inizio, 10, 1), date(anno_inizio, 10, 31)),
            11: (date(anno_inizio, 11, 1), date(anno_inizio, 11, 30)),
            12: (date(anno_inizio, 12, 1), date(anno_inizio, 12, 22)),  # vacanze di Natale
            1:  (date(anno_fine, 1, 8), date(anno_fine, 1, 31)),        # rientro ~8/1
            2:  (date(anno_fine, 2, 1), date(anno_fine, 2, 28)),
            3:  (date(anno_fine, 3, 1), date(anno_fine, 3, 31)),
            4:  (date(anno_fine, 4, 1), date(anno_fine, 4, 30)),
            5:  (date(anno_fine, 5, 1), date(anno_fine, 5, 31)),
            6:  (date(anno_fine, 6, 1), date(anno_fine, 6, 30)),        # infanzia fino al 30
        }
        # Giugno non-infanzia: finisce l'8
        FINESTRA_GIUGNO_ALTRI = (date(anno_fine, 6, 1), date(anno_fine, 6, 8))

        def conta_giorni(inizio, fine):
            """Conta giorni feriali (lun-ven) tra inizio e fine inclusi,
            escluse festivita' fisse e vacanze pasquali."""
            count = 0
            d = inizio
            while d <= fine:
                if (d.weekday() < 5
                        and (d.day, d.month) not in FESTIVITA_FISSE
                        and d not in vacanze_pasquali):
                    count += 1
                d += timedelta(days=1)
            return count

        mesi_calcolati = 0
        MESI_SCOLASTICI_LIST = [
            (9, anno_inizio), (10, anno_inizio), (11, anno_inizio), (12, anno_inizio),
            (1, anno_fine), (2, anno_fine), (3, anno_fine), (4, anno_fine), (5, anno_fine), (6, anno_fine)
        ]

        with db.get_db_context() as conn:
            cursor = conn.cursor()
            for mese, anno in MESI_SCOLASTICI_LIST:
                inizio, fine = FINESTRE[mese]
                giorni = conta_giorni(inizio, fine)

                # Solo a giugno calcoliamo anche la variante non-infanzia
                giorni_altri = None
                if mese == 6:
                    g_inizio, g_fine = FINESTRA_GIUGNO_ALTRI
                    giorni_altri = conta_giorni(g_inizio, g_fine)

                cursor.execute('''
                    INSERT OR REPLACE INTO calendario_scolastico
                    (anno_scolastico, mese, anno, giorni_lavorativi, giorni_lavorativi_altri)
                    VALUES (?, ?, ?, ?, ?)
                ''', (anno_scolastico, mese, anno, giorni, giorni_altri))
                mesi_calcolati += 1

        return jsonify({
            'success': True,
            'mesi_calcolati': mesi_calcolati
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/anni-scolastici', methods=['GET'])
def api_get_anni_scolastici():
    """Ottiene lista anni scolastici"""
    anni = db.get_anni_scolastici()
    return jsonify(anni)


@app.route('/api/anni-scolastici', methods=['POST'])
def api_create_anno_scolastico():
    """Crea nuovo anno scolastico"""
    data = request.json
    anno_inizio = data.get('anno_inizio')
    if not anno_inizio:
        return jsonify({'error': 'anno_inizio richiesto'}), 400

    anno_scolastico = db.create_anno_scolastico(anno_inizio)
    return jsonify({'anno_scolastico': anno_scolastico})





@app.route('/api/stats/utenti-da-completare/<int:anno>/<int:mese>')
def api_utenti_da_completare(anno, mese):
    """Ottiene gli utenti senza ore registrate per il mese specificato"""
    with db.get_db_context() as conn:
        cursor = conn.cursor()

        # Ottieni utenti attivi senza ore per questo mese
        cursor.execute('''
            SELECT u.id, u.nome, u.cognome, u.monte_ore_settimanale,
                   s.nome_completo as scuola, c.nome as commessa
            FROM utenti u
            JOIN scuole s ON u.scuola_id = s.id
            JOIN commesse c ON s.commessa_id = c.id
            LEFT JOIN rendicontazione r ON u.id = r.utente_id
                AND r.anno = ? AND r.mese = ?
            WHERE u.attivo = 1
                AND (r.ore_lavorate_60 IS NULL OR r.ore_lavorate_60 = 0)
            ORDER BY c.nome, s.nome_completo, u.cognome, u.nome
        ''', (anno, mese))

        utenti = []
        for row in cursor.fetchall():
            utenti.append({
                'id': row['id'],
                'nome': row['nome'],
                'cognome': row['cognome'],
                'nome_completo': f"{row['nome']} {row['cognome']}",
                'monte_ore': row['monte_ore_settimanale'],
                'scuola': row['scuola'],
                'commessa': row['commessa']
            })

    return jsonify(utenti)


@app.route('/api/stats')
def api_stats():
    """Statistiche generali"""
    with db.get_db_context() as conn:
        cursor = conn.cursor()

        # Conta utenti
        cursor.execute("SELECT COUNT(*) as count FROM utenti WHERE attivo = 1")
        num_utenti = cursor.fetchone()['count']

        # Conta scuole
        cursor.execute("SELECT COUNT(*) as count FROM scuole")
        num_scuole = cursor.fetchone()['count']

        # Conta utenti per commessa
        cursor.execute('''
            SELECT c.nome, COUNT(u.id) as count
            FROM commesse c
            LEFT JOIN scuole s ON c.id = s.commessa_id
            LEFT JOIN utenti u ON s.id = u.scuola_id AND u.attivo = 1
            GROUP BY c.id
        ''')
        utenti_per_commessa = {r['nome']: r['count'] for r in cursor.fetchall()}

    return jsonify({
        'num_utenti': num_utenti,
        'num_scuole': num_scuole,
        'utenti_per_commessa': utenti_per_commessa
    })


# ==================== UNDO ====================

@app.route('/api/undo', methods=['GET'])
def api_get_undo_stack():
    """Mostra le azioni annullabili (ora persistenti)"""
    undo_stack = db.get_undo_stack()
    actions = []
    for i, action in enumerate(undo_stack):
        desc = ''
        if action['type'] == 'delete_utente':
            u = action['data'].get('utente', {})
            desc = f'Eliminazione utente: {u.get("nome", "")} {u.get("cognome", "")}'
        elif action['type'] == 'update_utente':
            desc = f'Modifica utente ID {action["data"].get("id", "?")}'
        actions.append({
            'index': i,
            'tipo': action['type'],
            'descrizione': desc,
            'timestamp': action.get('timestamp', '')
        })
    return jsonify(actions[:10])


@app.route('/api/undo', methods=['POST'])
def api_undo():
    """Annulla l'ultima azione (persistente)"""
    action = db.pop_undo_action()

    if not action:
        return jsonify({'error': 'Nessuna azione da annullare'}), 400

    try:
        with db.get_db_context() as conn:
            cursor = conn.cursor()

            if action['type'] == 'delete_utente':
                u = action['data'].get('utente', {})
                # Ripristina utente
                cursor.execute('''
                    INSERT INTO utenti (id, scuola_id, nome, cognome, nome_puntato,
                        monte_ore_settimanale, attivo, lista_attesa, data_inserimento,
                        data_inizio, data_fine)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (u['id'], u['scuola_id'], u['nome'], u['cognome'],
                      u.get('nome_puntato', ''), u['monte_ore_settimanale'],
                      u.get('attivo', 1), u.get('lista_attesa'),
                      u.get('data_inserimento', datetime.now().isoformat()),
                      u.get('data_inizio'), u.get('data_fine')))

                # Ripristina rendicontazioni
                for r in action['data'].get('rendicontazioni', []):
                    cursor.execute('''
                        INSERT OR IGNORE INTO rendicontazione
                        (utente_id, anno, mese, ore_lavorate_60,
                         pasti, giorni_lavorativi, note, data_inserimento)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (r['utente_id'], r['anno'], r['mese'],
                          r.get('ore_lavorate_60'), r.get('pasti'),
                          r.get('giorni_lavorativi', 0), r.get('note'),
                          r.get('data_inserimento', datetime.now().isoformat())))

                # Ripristina note, documenti e assenze (presenti negli snapshot
                # creati da raccogli_snapshot_utente; assenti negli undo storici)
                for n in action['data'].get('note', []):
                    cursor.execute('''
                        INSERT OR IGNORE INTO note_utente
                        (id, utente_id, tipo, anno, mese, contenuto, priorita,
                         data_creazione, data_modifica)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (n.get('id'), n['utente_id'], n.get('tipo', 'generale'),
                          n.get('anno'), n.get('mese'), n.get('contenuto', ''),
                          n.get('priorita', 'normale'),
                          n.get('data_creazione', datetime.now().isoformat()),
                          n.get('data_modifica')))
                for doc in action['data'].get('documenti', []):
                    cursor.execute('''
                        INSERT OR IGNORE INTO documenti_utente
                        (id, utente_id, nome_file, nome_originale, tipo_documento,
                         descrizione, data_scadenza, data_caricamento, dimensione)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (doc.get('id'), doc['utente_id'], doc.get('nome_file', ''),
                          doc.get('nome_originale', ''), doc.get('tipo_documento', 'altro'),
                          doc.get('descrizione'), doc.get('data_scadenza'),
                          doc.get('data_caricamento', datetime.now().isoformat()),
                          doc.get('dimensione')))
                for a in action['data'].get('assenze', []):
                    cursor.execute('''
                        INSERT OR IGNORE INTO assenze
                        (id, utente_id, data_inizio, data_fine, tipo, motivazione,
                         note, data_registrazione)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (a.get('id'), a['utente_id'], a.get('data_inizio', ''),
                          a.get('data_fine'), a.get('tipo', 'altro'), a.get('motivazione'),
                          a.get('note'),
                          a.get('data_registrazione', datetime.now().isoformat())))

                nome = f"{u['nome']} {u.get('cognome', '')}"
                db.log_audit('undo', 'utente', u['id'], f'Ripristinato utente {nome}')
                return jsonify({'success': True, 'message': f'Utente {nome} ripristinato'})

            elif action['type'] == 'update_utente':
                old = action['data'].get('dati_precedenti', {})
                uid = action['data'].get('id')
                set_parts = []
                params = []
                for key in ['nome', 'cognome', 'monte_ore_settimanale', 'nome_puntato', 'lista_attesa']:
                    if key in old:
                        set_parts.append(f"{key} = ?")
                        params.append(old[key])
                if set_parts:
                    params.append(uid)
                    cursor.execute(f"UPDATE utenti SET {', '.join(set_parts)} WHERE id = ?", params)

                db.log_audit('undo', 'utente', uid, 'Ripristinati dati precedenti')
                return jsonify({'success': True, 'message': 'Modifica annullata'})

            return jsonify({'error': 'Tipo azione non supportato per undo'}), 400

    except Exception as e:
        logger.error(f"Errore undo: {e}")
        return jsonify({'error': str(e)}), 500


# ==================== CONFRONTO ANNUALE ====================

@app.route('/api/stats/confronto-annuale')
def api_confronto_annuale():
    """Confronta i dati di due anni scolastici"""
    anno_1 = request.args.get('anno_1')
    anno_2 = request.args.get('anno_2')

    if not anno_1 or not anno_2:
        return jsonify({'error': 'Specificare anno_1 e anno_2 (es: 2024-2025)'}), 400

    confronto = db.get_confronto_annuale(anno_1, anno_2)
    return jsonify({
        'anno_1': anno_1,
        'anno_2': anno_2,
        'confronto': confronto
    })


# ==================== STATISTICHE PER SCUOLA ====================

@app.route('/api/stats/per-scuola/<int:anno>/<int:mese>')
def api_stats_per_scuola(anno, mese):
    """Statistiche dettagliate per scuola, per grafici"""
    commessa = request.args.get('commessa')
    totali = db.get_totali_per_scuola(anno, mese, commessa)

    risultati = []
    for t in totali:
        risultati.append({
            'scuola': t['scuola'],
            'commessa': t['commessa'],
            'num_utenti': t['num_utenti'],
            'ore_erogate': round(t['ore_lavorate_60'] or 0, 2),
            'pasti': t['pasti'] or 0,
            'credito_debito': round(t['credito_debito'] or 0, 2),
            'imponibile': t.get('imponibile_100', 0),
            'totale': t.get('totale_100', 0)
        })

    return jsonify(risultati)


# ==================== PAGINA DETTAGLIO UTENTE ====================

@app.route('/utente/<int:utente_id>')
def utente_dettaglio_page(utente_id):
    """Pagina dettaglio utente"""
    return render_template('utente_dettaglio.html', utente_id=utente_id)


# ==================== CONFIG API ====================

# ==================== API STATISTICHE AVANZATE ====================

@app.route('/api/stats/heatmap/<anno_scolastico>')
def api_stats_heatmap(anno_scolastico):
    """Heatmap presenza mensile per utente"""
    try:
        commessa = request.args.get('commessa')
        limit = request.args.get('limit', 50, type=int)

        anni = anno_scolastico.split('-')
        anno_inizio = int(anni[0])
        anno_fine = int(anni[1])

        with db.get_db_context() as conn:
            # Ottieni utenti con JOIN per commessa e scuola
            query_utenti = """
                SELECT u.id, u.nome, u.cognome, cm.nome as commessa,
                       s.nome_completo as scuola, u.monte_ore_settimanale
                FROM utenti u
                JOIN scuole s ON u.scuola_id = s.id
                JOIN commesse cm ON s.commessa_id = cm.id
                WHERE u.attivo = 1
            """
            params = []
            if commessa:
                query_utenti += " AND cm.nome = ?"
                params.append(commessa)
            query_utenti += " ORDER BY u.cognome, u.nome LIMIT ?"
            params.append(limit)

            cursor = conn.execute(query_utenti, params)
            utenti = [dict(row) for row in cursor.fetchall()]

            # Mesi dell'anno scolastico
            mesi_scolastici = [
                {'mese': 9, 'anno': anno_inizio},
                {'mese': 10, 'anno': anno_inizio},
                {'mese': 11, 'anno': anno_inizio},
                {'mese': 12, 'anno': anno_inizio},
                {'mese': 1, 'anno': anno_fine},
                {'mese': 2, 'anno': anno_fine},
                {'mese': 3, 'anno': anno_fine},
                {'mese': 4, 'anno': anno_fine},
                {'mese': 5, 'anno': anno_fine},
                {'mese': 6, 'anno': anno_fine}
            ]

            # Pre-calcolo per evitare una query per ogni (utente x mese): variazioni
            # monte ore e giorni di calendario una volta per mese, e le ore erogate
            # di tutti gli utenti in un'unica query bulk.
            variazioni_per_mese = {}
            giorni_per_mese = {}
            for m in mesi_scolastici:
                variazioni_per_mese[(m['mese'], m['anno'])] = db.get_monte_ore_effettivo_bulk(m['anno'], m['mese'])
                giorni_per_mese[(m['mese'], m['anno'])] = db.get_calendario(anno_scolastico, m['mese'], m['anno'])

            ore_map = {}
            utente_ids = [u['id'] for u in utenti]
            if utente_ids:
                ph = ','.join('?' * len(utente_ids))
                cursor = conn.execute(f"""
                    SELECT utente_id, anno, mese, COALESCE(SUM(ore_lavorate_60), 0) as ore
                    FROM rendicontazione
                    WHERE utente_id IN ({ph})
                      AND ((anno = ? AND mese >= 9) OR (anno = ? AND mese <= 6))
                    GROUP BY utente_id, anno, mese
                """, utente_ids + [anno_inizio, anno_fine])
                for r in cursor.fetchall():
                    ore_map[(r['utente_id'], r['anno'], r['mese'])] = r['ore']

            # Componi la heatmap: nessuna query dentro il loop
            heatmap_data = []
            for utente in utenti:
                utente_data = {
                    'id': utente['id'],
                    'nome': f"{utente['nome']} {utente['cognome']}",
                    'commessa': utente['commessa'],
                    'monte_ore': utente['monte_ore_settimanale'],
                    'mesi': []
                }

                for m in mesi_scolastici:
                    ore = ore_map.get((utente['id'], m['anno'], m['mese']), 0)

                    # Monte ore effettivo (con variazione se presente)
                    var_mese = variazioni_per_mese.get((m['mese'], m['anno']), {})
                    monte_ore_eff = var_mese.get(utente['id'], utente['monte_ore_settimanale'] or 0)

                    # Ore previste per il mese (formula centralizzata)
                    giorni_lav = giorni_per_mese[(m['mese'], m['anno'])]
                    _, ore_previste_ridotte = db.calcola_media_prevista(monte_ore_eff, giorni_lav)

                    percentuale = (ore / ore_previste_ridotte * 100) if ore_previste_ridotte > 0 else 0

                    utente_data['mesi'].append({
                        'mese': m['mese'],
                        'anno': m['anno'],
                        'ore': round(ore, 2),
                        'ore_previste': round(ore_previste_ridotte, 2),
                        'percentuale': round(min(percentuale, 150), 1)  # Cap a 150%
                    })

                heatmap_data.append(utente_data)

            return jsonify({
                'heatmap': heatmap_data,
                'mesi': [{'mese': m['mese'], 'anno': m['anno']} for m in mesi_scolastici]
            })
    except Exception as e:
        logger.error(f"Errore heatmap: {e}")
        return jsonify({'error': 'Errore nel calcolo della heatmap'}), 500


@app.route('/api/stats/scuole-dettaglio')
def api_stats_scuole_dettaglio():
    """Statistiche dettagliate per scuola con filtri"""
    try:
        anno = request.args.get('anno', type=int)
        mese = request.args.get('mese', type=int)
        commessa = request.args.get('commessa')
        order_by = request.args.get('order_by', 'ore_erogate')
        order_dir = request.args.get('order_dir', 'desc')

        with db.get_db_context() as conn:
            # NOTA: utenti.scuola e utenti.commessa non esistono, facciamo JOIN
            query = """
                SELECT
                    s.nome_completo as scuola,
                    cm.nome as commessa,
                    COUNT(DISTINCT u.id) as num_utenti,
                    SUM(u.monte_ore_settimanale) as monte_ore_totale
            """

            if anno and mese:
                # Solo le ore in SQL: imponibile/totale-IVA derivati in Python da
                # config (mai costanti tariffa/IVA hardcoded nelle query).
                query += """,
                    COALESCE(SUM(r.ore_lavorate_60), 0) as ore_erogate
                """

            query += """
                FROM utenti u
                JOIN scuole s ON u.scuola_id = s.id
                JOIN commesse cm ON s.commessa_id = cm.id
            """

            if anno and mese:
                query += """
                    LEFT JOIN rendicontazione r ON u.id = r.utente_id
                        AND r.anno = ? AND r.mese = ?
                """

            query += " WHERE u.attivo = 1"

            params = []
            if anno and mese:
                params.extend([anno, mese])

            if commessa:
                query += " AND cm.nome = ?"
                params.append(commessa)

            query += " GROUP BY s.id, s.nome_completo, cm.nome"

            # Order by (usa gli alias definiti nella SELECT). 'ore_erogate' esiste
            # solo con anno/mese: altrimenti si ripiega su una colonna sempre presente
            # per non generare "no such column".
            valid_orders = {
                'num_utenti': 'num_utenti',
                'monte_ore_totale': 'monte_ore_totale',
                'scuola': 's.nome_completo'
            }
            if anno and mese:
                valid_orders['ore_erogate'] = 'ore_erogate'
            colonna_order = valid_orders.get(order_by, 'num_utenti')
            direction = 'DESC' if order_dir.lower() == 'desc' else 'ASC'
            query += f" ORDER BY {colonna_order} {direction}"

            cursor = conn.execute(query, params)
            scuole = [dict(row) for row in cursor.fetchall()]

            # Deriva imponibile e totale IVA-inclusa dalle ore, via config
            if anno and mese:
                for s in scuole:
                    imponibile, _iva, totale = config.calcola_fatturazione(s.get('ore_erogate', 0))
                    s['imponibile'] = imponibile
                    s['totale_iva'] = totale

            return jsonify({'scuole': scuole})
    except Exception as e:
        logger.error(f"Errore stats scuole: {e}")
        return jsonify({'error': 'Errore nel calcolo delle statistiche'}), 500


@app.route('/api/stats/validazione')
def api_stats_validazione():
    """Verifica anomalie nei dati e restituisce avvisi"""
    try:
        anno = request.args.get('anno', type=int)
        mese = request.args.get('mese', type=int)

        if not anno or not mese:
            now = datetime.now()
            anno = anno or now.year
            mese = mese or now.month

        anomalie = []

        with db.get_db_context() as conn:
            # 1. Utenti senza ore nel mese
            cursor = conn.execute("""
                SELECT u.id, u.nome, u.cognome, cm.nome as commessa,
                       s.nome_completo as scuola, u.monte_ore_settimanale
                FROM utenti u
                JOIN scuole s ON u.scuola_id = s.id
                JOIN commesse cm ON s.commessa_id = cm.id
                WHERE u.attivo = 1
                AND u.monte_ore_settimanale > 0
                AND u.id NOT IN (
                    SELECT DISTINCT utente_id FROM rendicontazione
                    WHERE anno = ? AND mese = ? AND ore_lavorate_60 > 0
                )
            """, [anno, mese])
            utenti_senza_ore = [dict(row) for row in cursor.fetchall()]

            if utenti_senza_ore:
                anomalie.append({
                    'tipo': 'warning',
                    'categoria': 'ore_mancanti',
                    'titolo': 'Utenti senza ore registrate',
                    'messaggio': f'{len(utenti_senza_ore)} utenti non hanno ore registrate questo mese',
                    'conteggio': len(utenti_senza_ore),
                    'dettagli': [{'id': u['id'], 'nome': f"{u['nome']} {u['cognome']}", 'commessa': u['commessa']}
                                for u in utenti_senza_ore[:10]]
                })

            # 2. Utenti con ore negative o anomale
            cursor = conn.execute("""
                SELECT u.id, u.nome, u.cognome, r.ore_lavorate_60, u.monte_ore_settimanale
                FROM rendicontazione r
                JOIN utenti u ON r.utente_id = u.id
                WHERE r.anno = ? AND r.mese = ?
                AND (r.ore_lavorate_60 < 0 OR r.ore_lavorate_60 > 200)
            """, [anno, mese])
            ore_anomale = [dict(row) for row in cursor.fetchall()]

            if ore_anomale:
                anomalie.append({
                    'tipo': 'danger',
                    'categoria': 'ore_anomale',
                    'titolo': 'Ore con valori anomali',
                    'messaggio': f'{len(ore_anomale)} registrazioni hanno valori sospetti',
                    'conteggio': len(ore_anomale),
                    'dettagli': [{'id': u['id'], 'nome': f"{u['nome']} {u['cognome']}", 'ore': u['ore_lavorate_60']}
                                for u in ore_anomale]
                })

            # 3. Utenti con monte ore = 0 ma con ore lavorate
            cursor = conn.execute("""
                SELECT u.id, u.nome, u.cognome, r.ore_lavorate_60
                FROM rendicontazione r
                JOIN utenti u ON r.utente_id = u.id
                WHERE r.anno = ? AND r.mese = ?
                AND u.monte_ore_settimanale = 0
                AND r.ore_lavorate_60 > 0
            """, [anno, mese])
            monte_ore_zero = [dict(row) for row in cursor.fetchall()]

            if monte_ore_zero:
                anomalie.append({
                    'tipo': 'info',
                    'categoria': 'monte_ore_zero',
                    'titolo': 'Utenti con monte ore non configurato',
                    'messaggio': f'{len(monte_ore_zero)} utenti hanno ore lavorate ma monte ore = 0',
                    'conteggio': len(monte_ore_zero),
                    'dettagli': [{'id': u['id'], 'nome': f"{u['nome']} {u['cognome']}", 'ore': u['ore_lavorate_60']}
                                for u in monte_ore_zero]
                })

            # 4. Utenti con ore erogate molto diverse dalle previste (>50% differenza)
            cursor = conn.execute("""
                SELECT
                    u.id, u.nome, u.cognome, u.monte_ore_settimanale,
                    COALESCE(SUM(r.ore_lavorate_60), 0) as ore_erogate
                FROM utenti u
                LEFT JOIN rendicontazione r ON u.id = r.utente_id AND r.anno = ? AND r.mese = ?
                WHERE u.attivo = 1 AND u.monte_ore_settimanale > 0
                GROUP BY u.id
            """, [anno, mese])

            # Determina anno scolastico
            anno_scolastico_validazione = f"{anno}-{anno+1}" if mese >= 9 else f"{anno-1}-{anno}"
            giorni_lav = db.get_calendario(anno_scolastico_validazione, mese, anno)
            differenze_anomale = []
            variazioni_anomalie = db.get_monte_ore_effettivo_bulk(anno, mese)

            for row in cursor.fetchall():
                monte_ore_eff = variazioni_anomalie.get(row['id'], row['monte_ore_settimanale'])
                _, ore_previste_ridotte = db.calcola_media_prevista(monte_ore_eff, giorni_lav)
                ore_erogate = row['ore_erogate']

                if ore_previste_ridotte > 0:
                    diff_percentuale = abs(ore_erogate - ore_previste_ridotte) / ore_previste_ridotte * 100
                    if diff_percentuale > config.SOGLIA_ANOMALIA_PERCENTUALE and ore_erogate > 0:
                        differenze_anomale.append({
                            'id': row['id'],
                            'nome': f"{row['nome']} {row['cognome']}",
                            'ore_erogate': round(ore_erogate, 2),
                            'ore_previste': round(ore_previste_ridotte, 2),
                            'diff_percentuale': round(diff_percentuale, 1)
                        })

            if differenze_anomale:
                anomalie.append({
                    'tipo': 'warning',
                    'categoria': 'differenze_elevate',
                    'titolo': 'Differenze significative ore',
                    'messaggio': f'{len(differenze_anomale)} utenti con differenza >50% tra ore erogate e previste',
                    'conteggio': len(differenze_anomale),
                    'dettagli': differenze_anomale[:10]
                })

            # 5. Commesse senza dati nel mese
            cursor = conn.execute("""
                SELECT c.nome, COUNT(u.id) as num_utenti
                FROM commesse c
                LEFT JOIN scuole s ON s.commessa_id = c.id
                LEFT JOIN utenti u ON u.scuola_id = s.id AND u.attivo = 1
                WHERE c.attiva = 1
                GROUP BY c.id
                HAVING num_utenti = 0
            """)
            commesse_vuote = [dict(row) for row in cursor.fetchall()]

            if commesse_vuote:
                anomalie.append({
                    'tipo': 'info',
                    'categoria': 'commesse_vuote',
                    'titolo': 'Commesse senza utenti',
                    'messaggio': f'{len(commesse_vuote)} commesse attive non hanno utenti associati',
                    'conteggio': len(commesse_vuote),
                    'dettagli': [{'nome': c['nome']} for c in commesse_vuote]
                })

        # Riepilogo
        riepilogo = {
            'totale_anomalie': len(anomalie),
            'critiche': len([a for a in anomalie if a['tipo'] == 'danger']),
            'avvisi': len([a for a in anomalie if a['tipo'] == 'warning']),
            'info': len([a for a in anomalie if a['tipo'] == 'info'])
        }

        return jsonify({
            'anomalie': anomalie,
            'riepilogo': riepilogo,
            'periodo': {'anno': anno, 'mese': mese}
        })
    except Exception as e:
        logger.error(f"Errore validazione: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/stats/confronto-annuale-dettaglio')
def api_stats_confronto_annuale_dettaglio():
    """Confronto dettagliato tra anni scolastici"""
    try:
        mese = request.args.get('mese', type=int)
        commessa = request.args.get('commessa')

        if not mese:
            mese = datetime.now().month

        with db.get_db_context() as conn:
            # Ottieni anni scolastici disponibili
            cursor = conn.execute("SELECT DISTINCT anno_scolastico FROM calendario_scolastico ORDER BY anno_scolastico DESC LIMIT 5")
            anni = [row['anno_scolastico'] for row in cursor.fetchall()]

            risultati = []
            for anno_scolastico in anni:
                anni_parts = anno_scolastico.split('-')
                anno_effettivo = int(anni_parts[0]) if mese >= 9 else int(anni_parts[1])

                query = """
                    SELECT
                        COUNT(DISTINCT r.utente_id) as num_utenti,
                        COALESCE(SUM(r.ore_lavorate_60), 0) as ore_erogate
                    FROM rendicontazione r
                    JOIN utenti u ON r.utente_id = u.id
                """
                params = [anno_effettivo, mese]

                if commessa:
                    query += """
                    JOIN scuole s ON u.scuola_id = s.id
                    JOIN commesse cm ON s.commessa_id = cm.id
                    WHERE r.anno = ? AND r.mese = ? AND cm.nome = ?
                    """
                    params.append(commessa)
                else:
                    query += " WHERE r.anno = ? AND r.mese = ?"

                cursor = conn.execute(query, params)
                dati = cursor.fetchone()

                risultati.append({
                    'anno_scolastico': anno_scolastico,
                    'anno': anno_effettivo,
                    'mese': mese,
                    'num_utenti': dati['num_utenti'],
                    'ore_erogate': round(dati['ore_erogate'], 2),
                    'imponibile': config.calcola_fatturazione(dati['ore_erogate'])[0]
                })

            # Calcola variazioni percentuali
            for i in range(len(risultati) - 1):
                if risultati[i + 1]['ore_erogate'] > 0:
                    variazione = ((risultati[i]['ore_erogate'] - risultati[i + 1]['ore_erogate'])
                                 / risultati[i + 1]['ore_erogate'] * 100)
                    risultati[i]['variazione'] = round(variazione, 1)
                else:
                    risultati[i]['variazione'] = None

            if risultati:
                risultati[-1]['variazione'] = None

            return jsonify({'confronto': risultati})
    except Exception as e:
        logger.error(f"Errore confronto annuale: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/utenti/filtrati')
def api_utenti_filtrati():
    """API per ottenere lista utenti con filtri avanzati"""
    try:
        commessa = request.args.get('commessa')
        scuola = request.args.get('scuola')
        search = request.args.get('search', '').strip()

        with db.get_db_context() as conn:
            # NOTA: utenti.commessa, utenti.scuola non esistono, usiamo JOIN.
            query = """
                SELECT u.id, u.nome, u.cognome,
                       cm.nome as commessa,
                       s.nome_completo as scuola,
                       u.monte_ore_settimanale
                FROM utenti u
                JOIN scuole s ON u.scuola_id = s.id
                JOIN commesse cm ON s.commessa_id = cm.id
                WHERE u.attivo = 1
            """
            params = []

            if commessa:
                query += " AND cm.nome = ?"
                params.append(commessa)

            if scuola:
                query += " AND s.nome_completo LIKE ?"
                params.append(f"%{scuola}%")

            if search:
                query += " AND (u.nome LIKE ? OR u.cognome LIKE ?)"
                params.extend([f"%{search}%", f"%{search}%"])

            query += " ORDER BY u.cognome, u.nome"

            cursor = conn.execute(query, params)
            utenti = [dict(row) for row in cursor.fetchall()]

            return jsonify(utenti)
    except Exception as e:
        logger.error(f"Errore utenti filtrati: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/scuole/lista')
def api_scuole_lista():
    """Lista scuole per filtri dropdown"""
    try:
        commessa = request.args.get('commessa')

        with db.get_db_context() as conn:
            # NOTA: utenti.scuola/utenti.commessa non esistono, facciamo JOIN
            query = """
                SELECT s.nome_completo as scuola,
                       cm.nome as commessa,
                       COUNT(u.id) as num_utenti
                FROM scuole s
                JOIN commesse cm ON s.commessa_id = cm.id
                LEFT JOIN utenti u ON u.scuola_id = s.id AND u.attivo = 1
                WHERE s.nome_completo IS NOT NULL AND s.nome_completo != ''
            """
            params = []

            if commessa:
                query += " AND cm.nome = ?"
                params.append(commessa)

            query += " GROUP BY s.id, s.nome_completo, cm.nome ORDER BY s.nome_completo"

            cursor = conn.execute(query, params)
            scuole = [dict(row) for row in cursor.fetchall()]

            return jsonify(scuole)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/config')
def api_get_config():
    """Ritorna la configurazione corrente (parametri di calcolo)"""
    return jsonify({
        'tariffa_oraria': config.TARIFFA_ORARIA,
        'iva_percentuale': config.IVA_PERCENTUALE,
        'tasso_assenza': config.TASSO_ASSENZA,
        'coefficiente_giornaliero': config.COEFFICIENTE_GIORNALIERO,
        'max_ore_settimanali': config.MAX_ORE_SETTIMANALI,
        'max_pasti_mensili': config.MAX_PASTI_MENSILI
    })


def open_browser():
    """Apre il browser dopo un breve ritardo"""
    import webbrowser
    import time
    time.sleep(1.5)
    webbrowser.open('http://localhost:5000')


if __name__ == '__main__':
    import sys
    import threading

    logger.info("=" * 50)
    logger.info("  GESTIONALE OEPAC - Sistema di Rendicontazione")
    logger.info("=" * 50)
    logger.info("  Avvio server su http://localhost:5000")

    if getattr(sys, 'frozen', False):
        logger.info("  Apertura browser in corso...")
        threading.Thread(target=open_browser, daemon=True).start()
        app.run(host='127.0.0.1', port=5000, debug=False, use_reloader=False)
    else:
        logger.info("  Premi Ctrl+C per terminare")
        # Default SICURI anche in esecuzione da sorgente: nessuna esposizione di
        # rete (solo loopback) e nessun debugger interattivo Werkzeug (che
        # consentirebbe esecuzione di codice arbitrario a chi raggiunge la porta).
        # Per lo sviluppo si possono attivare esplicitamente via variabili
        # d'ambiente: FLASK_DEBUG=1 abilita debugger+reloader, OEPAC_HOST cambia il bind.
        host = os.environ.get('OEPAC_HOST', '127.0.0.1')
        debug = os.environ.get('FLASK_DEBUG', '0') == '1'
        app.run(host=host, port=5000, debug=debug, use_reloader=debug)
