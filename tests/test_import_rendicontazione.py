"""Test del parsing e abbinamento per l'import delle rendicontazioni mensili."""

import io

import openpyxl
import pytest

import import_rendicontazione as imp


# ---------- helper parsing valori ----------

def test_parse_ore_numeri_e_stringhe():
    assert imp.parse_ore(45) == 45.0
    assert imp.parse_ore(18.17) == 18.17
    assert imp.parse_ore(None) is None
    assert imp.parse_ore('') is None
    assert imp.parse_ore('1,30') == 1.30          # virgola -> punto (decimale)
    assert imp.parse_ore('1:30') == pytest.approx(1.5)   # sessagesimale -> decimale
    assert imp.parse_ore('abc') is None


def test_parse_pasti():
    assert imp.parse_pasti(None) == 0
    assert imp.parse_pasti('') == 0
    assert imp.parse_pasti(7) == 7
    assert imp.parse_pasti(7.0) == 7
    assert imp.parse_pasti('3') == 3


def test_parse_mese_anno():
    assert imp.parse_mese_anno('ESTRATTO SETTEMBRE 2025') == (9, 2025)
    assert imp.parse_mese_anno('Estratto Dicembre') == (12, None)
    assert imp.parse_mese_anno('Foglio1') == (None, None)


def test_chiave_nome_ordine_e_spazi():
    # Indipendente da ordine e spazi multipli / accenti
    assert imp.chiave_nome('Mura Filippo') == imp.chiave_nome('Filippo Mura')
    assert imp.chiave_nome('Tocca  Giuliano') == imp.chiave_nome('Tocca Giuliano')
    assert imp.chiave_nome('Niccolò Rossi') == imp.chiave_nome('NICCOLO ROSSI')


# ---------- matching ----------

def _utenti_demo():
    return [
        {'id': 1, 'nome': 'Filippo', 'cognome': 'Mura', 'scuola': '/IC PALOMBINI/Secondaria/VILLARI'},
        {'id': 2, 'nome': 'Mario', 'cognome': 'Rossi', 'scuola': '/IC LAPARELLI/Primaria/REY'},
        {'id': 3, 'nome': 'Mario', 'cognome': 'Rossi', 'scuola': '/IC SALACONE/Secondaria/ROSA PARKS'},
    ]


def test_match_singolo():
    index = imp.build_index(_utenti_demo())
    u, stato = imp.match_riga({'nome_completo': 'Mura Filippo', 'scuola': '/IC PALOMBINI/...'}, index)
    assert stato == 'match' and u['id'] == 1


def test_match_non_trovato():
    index = imp.build_index(_utenti_demo())
    u, stato = imp.match_riga({'nome_completo': 'Bianchi Anna', 'scuola': '/IC X'}, index)
    assert stato == 'non_trovato' and u is None


def test_match_omonimi_disambigua_per_scuola():
    index = imp.build_index(_utenti_demo())
    u, stato = imp.match_riga(
        {'nome_completo': 'Rossi Mario', 'scuola': '/IC SALACONE/Secondaria/ROSA PARKS - Via Guattari'},
        index)
    assert stato == 'match' and u['id'] == 3


def test_match_omonimi_ambiguo_senza_scuola():
    index = imp.build_index(_utenti_demo())
    u, stato = imp.match_riga({'nome_completo': 'Rossi Mario', 'scuola': '/IC IGNOTO/Primaria'}, index)
    assert stato == 'ambiguo' and u is None


# ---------- workbook completo (in memoria) ----------

def _crea_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Estratto Ottobre'
    ws['A1'] = 'ESTRATTO OTTOBRE 2025'
    ws.append([])  # riga 2 vuota lasciata da A1? no: ricreiamo con append controllato
    # Ricostruiamo: riga1 titolo (gia' impostata), riga2 header, dati
    ws['A2'] = 'Commessa'
    ws['B2'] = 'Scuola'
    ws['C2'] = 'Utenti'
    ws['D2'] = "Totale ore lavorate in 100'"
    ws['E2'] = 'Pasti'
    ws.append(['OEPA IV', '/IC PALOMBINI/Secondaria/VILLARI', 'Mura Filippo', 28, 1])
    ws.append(['OEPA V', '/IC X/Primaria', 'Bianchi Anna', 30, None])
    ws.append(['Totale', None, None, 58, 1])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_parse_workbook_e_analizza():
    fogli = imp.parse_workbook(_crea_workbook())
    assert len(fogli) == 1
    f = fogli[0]
    assert f['mese'] == 10 and f['anno'] == 2025 and f['riconosciuto']
    # La riga "Totale" deve essere esclusa -> restano 2 righe dati
    assert len(f['righe']) == 2

    utenti = [{'id': 1, 'nome': 'Filippo', 'cognome': 'Mura', 'scuola': '/IC PALOMBINI/Secondaria/VILLARI'}]
    analisi = imp.analizza(fogli, utenti)[0]
    assert len(analisi['match']) == 1
    assert analisi['match'][0]['utente_id'] == 1
    assert analisi['match'][0]['ore'] == 28
    assert analisi['match'][0]['pasti'] == 1
    assert len(analisi['non_trovati']) == 1  # Bianchi Anna non in anagrafica
