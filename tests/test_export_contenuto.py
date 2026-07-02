"""Verifica del CONTENUTO degli export: i numeri nelle celle devono coincidere
con l'helper unico di fatturazione (config.calcola_fatturazione).
"""
import io

from openpyxl import load_workbook

import config


def _set_ore(db, uid, anno, mese, ore):
    db.get_or_create_rendicontazione(uid, anno, mese)
    db.update_rendicontazione(uid, anno, mese, ore_lavorate=ore)


def test_municipale_celle_coerenti_con_fatturazione(client, db_mod):
    """Nel riepilogo municipale imponibile/totale di ogni riga devono derivare
    dalle ore con l'helper unico (non da altri calcoli)."""
    db = db_mod
    db.create_commessa('OEPAC CELLE')
    sid = db.get_or_create_scuola('OEPAC CELLE', 'IC Celle - Primaria')
    u1 = db.get_or_create_utente(sid, 'Cella', 'Uno', 15)
    u2 = db.get_or_create_utente(sid, 'Cella', 'Due', 20)
    _set_ore(db, u1, 2026, 4, 61.5)
    _set_ore(db, u2, 2026, 4, 44.25)

    r = client.get('/api/export/municipale/2026/4?commessa=OEPAC%20CELLE')
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.data), data_only=True)
    ws = wb['Riepilogo Municipale']

    # layout reale: [scuola, ore60 HH:MM, tariffa, importo60, ore100, tariffa, importo100]
    righe = list(ws.iter_rows(values_only=True))
    riga = next((row for row in righe if row and row[0] == 'IC Celle - Primaria'), None)
    assert riga, 'riga della scuola di test non trovata nel foglio'

    ore_100 = riga[4]
    assert ore_100 == 61.5 + 44.25
    assert riga[2] == config.TARIFFA_ORARIA            # tariffa nella cella
    assert round(riga[6], 2) == round(ore_100 * config.TARIFFA_ORARIA, 2)

    # riepilogo fatturazione: imponibile/IVA/totale coerenti con l'helper unico
    imponibile, iva, totale = config.calcola_fatturazione(ore_100)
    valori = [round(c, 2) for row in righe for c in row
              if isinstance(c, (int, float)) and c is not None]
    for atteso in (imponibile, totale):
        assert atteso in valori, f'{atteso} assente dal riepilogo fatturazione'
